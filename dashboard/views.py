import io
import os
import re
import csv
import json
import random
import requests
import tempfile
import matplotlib
from collections import Counter
from datetime import datetime, timedelta, timezone as dt_tz
from django.conf import settings
from django.contrib import messages
import base64
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.mail import EmailMessage, send_mail
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference, LineChart
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from .models import UserLogin  # adjust if your model is elsewhere
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
)
from .models import ScheduledReport

# add at top (imports)
from django.utils import timezone
from zoneinfo import ZoneInfo
IST = ZoneInfo("Asia/Kolkata")

# ===========================
# imports used in this file
# ===========================
from datetime import datetime, timedelta, timezone as py_timezone, date
import numpy as np
matplotlib.use("Agg")  # ensure headless
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import PercentFormatter

# openpyxl
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# reportlab
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

import logging
logger = logging.getLogger(__name__)

###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################

# ====================================================================================
# AUTH MODULE (Signup / Login / Logout / OTP Reset)
# ====================================================================================

otp_store = {}  # temporary OTP memory


def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email")
        if not User.objects.filter(email=email).exists():
            messages.error(request, "Please enter your registered email address.")
            return redirect("forgot_password")

        user = User.objects.get(email=email)
        otp = random.randint(100000, 999999)
        otp_store[email] = otp

        send_mail(
            "Your KARVTECH Password Reset OTP",
            f"Hello {user.username},\n\nYour OTP is: {otp}\nValid for 10 minutes.\n\n-KARVTECH Team",
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )

        request.session["reset_email"] = email
        messages.success(request, "OTP sent to your email.")
        return redirect("verify_otp")

    return render(request, "forgot_password.html")


def verify_otp(request):
    email = request.session.get("reset_email")
    if not email:
        return redirect("forgot_password")

    if request.method == "POST":
        entered = request.POST.get("otp")
        actual = otp_store.get(email)
        if actual and str(entered) == str(actual):
            messages.success(request, "OTP verified. Reset password.")
            return redirect("reset_password")
        messages.error(request, "Invalid OTP.")

    return render(request, "verify_otp.html")


def reset_password(request):
    email = request.session.get("reset_email")
    if not email:
        return redirect("forgot_password")

    if request.method == "POST":
        pwd = request.POST.get("password")
        cpwd = request.POST.get("confirm_password")

        if pwd != cpwd:
            messages.error(request, "Passwords do not match.")
            return redirect("reset_password")

        user = User.objects.get(email=email)
        user.set_password(pwd)
        user.save()

        otp_store.pop(email, None)
        request.session.pop("reset_email", None)

        messages.success(request, "Password reset successfully!")
        return redirect("login")

    return render(request, "reset_password.html")


def signup_view(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]
        password = request.POST["password"]
        confirm = request.POST["confirm_password"]

        if password != confirm:
            messages.error(request, "Passwords do not match!")
            return redirect("signup")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists!")
            return redirect("signup")

        User.objects.create_user(username=username, email=email, password=password)
        messages.success(request, "Account created! Please log in.")
        return redirect("login")

    return render(request, "signup.html")


def login_view(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            messages.success(request, f"Welcome {username}!")
            return redirect("index")

        messages.error(request, "Invalid username/password.")
        return redirect("login")

    return render(request, "login.html")


def logout_view(request):
    logout(request)
    messages.info(request, "You have logged out.")
    storage = messages.get_messages(request)
    storage.used = True
    return redirect("login")


###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################


@login_required(login_url='login')
def index(request):
    return render(request, "index.html")

###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################

# ====================================================================================
# PROBLEM ANALYSIS REPORT  -  IST TIME UTILITIES + DYNATRACE HELPERS
# ====================================================================================

def convert_to_ist(epoch_ms):
    """Convert epoch ms to IST human-readable format."""
    if epoch_ms in (-1, None):
        return "Ongoing"

    try:
        IST = dt_tz(timedelta(hours=5, minutes=30))
        utc_dt = datetime.fromtimestamp(epoch_ms / 1000, tz=dt_tz.utc)
        return utc_dt.astimezone(IST).strftime("%d-%b-%Y %I:%M:%S %p")
    except:
        return "Invalid"


def _normalize_tenant(tenant_url):
    tenant_url = (tenant_url or "").strip().rstrip("/")
    if "apps.dynatrace.com" in tenant_url:
        tenant_url = tenant_url.replace("apps.dynatrace.com", "live.dynatrace.com")
    return tenant_url


def _fetch_management_zones(tenant_url, token):
    url = f"{tenant_url}/api/config/v1/managementZones"
    headers = {"Authorization": f"Api-Token {token}", "Accept": "application/json"}

    r = requests.get(url, headers=headers, timeout=60)
    r.raise_for_status()

    zones = r.json().get("values", [])
    return [{"id": z.get("id"), "name": z.get("name")} for z in zones]


def fetch_all_problems(tenant_url, token, start_time, end_time):
    """Fetch all problems with pagination."""
    all_data = []
    headers = {"Authorization": f"Api-Token {token}"}
    params = {
        "from": start_time,
        "to": end_time,
        "status": "OPEN,CLOSED",
        "pageSize": 100
    }

    url = f"{tenant_url}/api/v2/problems"

    while True:
        r = requests.get(url, headers=headers, params=params, timeout=60)
        r.raise_for_status()
        data = r.json()

        all_data.extend(data.get("problems", []))

        next_key = data.get("nextPageKey")
        if not next_key:
            break

        params = {"nextPageKey": next_key}

    return all_data


def _get_timeframe_range(timeframe):
    """Returns (start_utc, end_utc) in RFC3339 format."""
    IST = dt_tz(timedelta(hours=5, minutes=30))
    now_utc = datetime.now(dt_tz.utc)
    now_ist = now_utc.astimezone(IST)

    if timeframe == "1h":
        start_utc = now_utc - timedelta(hours=1)
    elif timeframe == "24h":
        start_utc = now_utc - timedelta(hours=24)
    elif timeframe == "today":
        start_ist = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
        start_utc = start_ist.astimezone(dt_tz.utc)
    elif timeframe == "yesterday":
        start_ist = (now_ist - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_ist = start_ist.replace(hour=23, minute=59, second=59)
        start_utc = start_ist.astimezone(dt_tz.utc)
        now_utc = end_ist.astimezone(dt_tz.utc)
    elif timeframe == "7d":
        start_ist = now_ist - timedelta(days=7)
        start_ist = start_ist.replace(hour=0, minute=0, second=0, microsecond=0)
        start_utc = start_ist.astimezone(dt_tz.utc)
    else:
        start_utc = now_utc - timedelta(hours=24)

    return (
        start_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
        now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
    )


# ====================================================================================
# PROBLEM DATASET BUILDER
# ====================================================================================

def _build_problem_dataset(tenant_url, token, timeframe, management_zone):
    tenant_url = _normalize_tenant(tenant_url)
    start, end = _get_timeframe_range(timeframe)
    problems = fetch_all_problems(tenant_url, token, start, end)

    # When user selects a specific zone (ID)
    # When user selects a specific zone (ID or Name)
    if management_zone and str(management_zone).lower() != "all":
        mz = str(management_zone).strip()

        problems = [
            p for p in problems
            if any(
                str(z.get("id")) == mz or str(z.get("name")) == mz
                for z in p.get("managementZones", [])
            )
        ]


    return problems


# ====================================================================================
# REPORT RENDERING ENGINE — SINGLE SOURCE OF TRUTH
# CSV / EXCEL / PDF / PPTX
# ====================================================================================

def _render_problem_report_bytes(problem_list, fmt, management_zone=""):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from collections import Counter, defaultdict

    fmt = (fmt or "").strip().lower()
    problem_list = problem_list or []

    # -------------------------------------------------------------------------
    # COMMON DATA
    # -------------------------------------------------------------------------
    all_rows = []
    entity_map = []  
    entity_event_map = []

    title_counts = Counter()
    entity_severity_counts = Counter()
    infra_map = Counter()
    service_map = Counter()

    for p in problem_list:
        entity_type, entity_name, root_cause_id = "", "", ""
        # Top 10 Category table
        title = p.get("title", "")
        if title:
            title_counts[title] += 1

                # Affected entity info
        if p.get("affectedEntities"):
            e = p["affectedEntities"][0]
            entity_type = e.get("entityType", "")
            entity_name = e.get("name", "")
            entity_id = e.get("entityId", "")
        else:
            entity_id = ""

        # Entity + Severity table
        if p.get("affectedEntities"):
            en = p["affectedEntities"][0].get("name", "")
            sev = p.get("severityLevel", "UNKNOWN")
            if en:
                entity_severity_counts[(en, sev)] += 1

        # Infrastructure table
        if p.get("impactLevel") == "INFRASTRUCTURE":
            en = p.get("affectedEntities", [{}])[0].get("name", "")
            sev = p.get("severityLevel", "UNKNOWN")
            if en:
                infra_map[(en, sev)] += 1

        # Services table
        if p.get("impactLevel") == "SERVICES":
            en = p.get("affectedEntities", [{}])[0].get("name", "")
            sev = p.get("severityLevel", "UNKNOWN")
            if en:
                service_map[(en, sev)] += 1

        # ————— Detailed Problems Table —————
        zones = ", ".join([z.get("name", "") for z in p.get("managementZones", [])])
        ename = p.get("affectedEntities", [{}])[0].get("name", "")
        etype = p.get("affectedEntities", [{}])[0].get("entityType", "")

        all_rows.append([
            str(p.get("displayId", "")),
            title,
            str(p.get("impactLevel", "")),
            str(p.get("severityLevel", "")),
            zones,
            etype,
            ename,
            str(convert_to_ist(p.get("startTime"))),
            str(convert_to_ist(p.get("endTime"))),
        ])
                # Entity → event mapping (for Excel & PDF)
        if entity_name:
            event_name = p.get("title", "")
            entity_event_map.append((entity_name, event_name))

        # Entity → entityId → rootCause (PDF)
        entity_map.append((entity_name, entity_id, root_cause_id))

    # ✅ GLOBAL COUNTERS
    title_counts = Counter([r[1] for r in all_rows if r[1]])
    entity_counter = Counter([name for name, evt in entity_event_map if name])

    # ===================================================================================
    # ✅ EXCEL FORMAT
    # ===================================================================================
    if fmt == "excel":
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Problem Analysis Report"

        # Styles
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=16, bold=True)
        sub_font = Font(size=11, bold=True)
        center_align = Alignment(horizontal="center")
        left_align = Alignment(horizontal="left")

        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # Helper
        def apply_border(ws, rng):
            for row in ws[rng]:
                for cell in row:
                    cell.border = thin_border

        # Title
        sheet1.merge_cells("A1:K4")
        sheet1["A1"] = f"Problem Analysis Report - {management_zone}"
        sheet1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet1["A1"].font = title_font
        sheet1["A1"].alignment = center_align


        sheet1.merge_cells("B6:F6")
        sheet1["B6"] = "Prepared by - KarvTech Team"
        sheet1["B6"].font = sub_font

        sheet1.merge_cells("B7:F7")
        sheet1["B7"] = f"Date - {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
        sheet1["B7"].font = sub_font

        start_row = 10

        # ✅ TOP EVENTS TABLE
        sheet1.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)
        c = sheet1.cell(row=start_row, column=2)
        c.value = "Top Problems by Event"
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align

        sheet1["B11"], sheet1["C11"] = "Event Name", "Count"
        for col in ("B11", "C11"):
            sheet1[col].fill = header_fill
            sheet1[col].font = header_font
            sheet1[col].alignment = center_align

        r = 12
        for title, cnt in title_counts.most_common(20):
            sheet1.cell(row=r, column=2, value=title)
            sheet1.cell(row=r, column=3, value=cnt)
            r += 1

        apply_border(sheet1, f"B11:C{r-1}")

        # PIE CHART
        pie = PieChart()
        pie.title = "Top Problems by Event"
        chart_data = Reference(sheet1, min_col=3, min_row=11, max_row=r - 1)
        labels = Reference(sheet1, min_col=2, min_row=12, max_row=r - 1)
        pie.add_data(chart_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width, pie.height = 11, 7

        sheet1.column_dimensions["E"].width = 25
        sheet1.merge_cells("E10:E20")
        sheet1.add_chart(pie, "E10")

        # ✅ PROBLEMATIC ENTITIES (Excel)
        start_col = 10
        sheet1.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
        cell = sheet1.cell(row=start_row, column=start_col)
        cell.value = "Problematic Entities"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

        headers = ["Entity Name", "Count", "Event Name"]
        for i, h in enumerate(headers):
            c = sheet1.cell(row=start_row + 1, column=start_col + i)
            c.value = h
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align

        rr = start_row + 2
        for name, cnt in entity_counter.most_common(40):
            evt = next((evt for en, evt in entity_event_map if en == name), "")
            sheet1.cell(row=rr, column=start_col, value=name)
            sheet1.cell(row=rr, column=start_col + 1, value=cnt)
            sheet1.cell(row=rr, column=start_col + 2, value=evt)
            rr += 1

        apply_border(sheet1, f"J11:L{rr-1}")

        # ✅ SHEET 2: Detailed
        sheet2 = wb.create_sheet("Detailed Problems")
        headers = [
            "Display ID", "Title", "Impact Level", "Severity Level",
            "Management Zones", "Entity Type", "Affected Entity Name",
            "Start (IST)", "End (IST)"
        ]
        sheet2.append(headers)

        for c in sheet2[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align

        for row in all_rows:
            sheet2.append(row)

        last = sheet2.max_row
        last_col = get_column_letter(sheet2.max_column)
        apply_border(sheet2, f"A1:{last_col}{last}")


        # ---------------------------------------------------------------------
        # SHEET 3 : Summary Tables
        # ---------------------------------------------------------------------
        sheet3 = wb.create_sheet("Summary Tables")
        row = 1

        # =========================================================
        # TABLE 1 – TOP 10 CATEGORIES
        # =========================================================
        sheet3["A1"] = "Top 10 Problems by Categories"
        sheet3["A1"].font = title_font
        row = 3

        sheet3["A3"], sheet3["B3"] = "Category", "Count"
        for col in ("A3", "B3"):
            sheet3[col].fill = header_fill
            sheet3[col].font = header_font
            sheet3[col].alignment = center_align

        r = 4
        for name, cnt in title_counts.most_common(10):
            sheet3.cell(r, 1).value = name
            sheet3.cell(r, 2).value = cnt
            r += 1

        apply_border(sheet3, f"A3:B{r-1}")
        row = r + 2

        # =========================================================
        # TABLE 2 – TOP 10 IMPACTED ENTITIES
        # =========================================================
        sheet3[f"A{row}"] = "Top 10 Problems by Impacted Entities"
        sheet3[f"A{row}"].font = title_font
        row += 2

        sheet3[f"A{row}"], sheet3[f"B{row}"], sheet3[f"C{row}"] = \
            "Entity", "Severity Level", "Count"

        for col in ("A", "B", "C"):
            c = sheet3[f"{col}{row}"]
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align

        row += 1
        start = row

        for (entity, sev), cnt in entity_severity_counts.most_common(10):
            sheet3.cell(row, 1).value = entity
            sheet3.cell(row, 2).value = sev
            sheet3.cell(row, 3).value = cnt
            row += 1

        apply_border(sheet3, f"A{start-1}:C{row-1}")
        row += 2

        # =========================================================
        # TABLE 3 – TOP 10 INFRASTRUCTURE PROBLEMS
        # =========================================================
        sheet3[f"A{row}"] = "Top 10 Infrastructure Problems"
        sheet3[f"A{row}"].font = title_font
        row += 2

        sheet3[f"A{row}"], sheet3[f"B{row}"], sheet3[f"C{row}"] = \
            "Infrastructure", "Severity Level", "Count"

        for col in ("A", "B", "C"):
            c = sheet3[f"{col}{row}"]
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align

        row += 1
        start = row

        for (infra, sev), cnt in infra_map.most_common(10):
            sheet3.cell(row, 1).value = infra
            sheet3.cell(row, 2).value = sev
            sheet3.cell(row, 3).value = cnt
            row += 1

        apply_border(sheet3, f"A{start-1}:C{row-1}")
        row += 2

        # =========================================================
        # TABLE 4 – TOP 10 SERVICES PROBLEMS
        # =========================================================
        sheet3[f"A{row}"] = "Top 10 Services Problems"
        sheet3[f"A{row}"].font = title_font
        row += 2

        sheet3[f"A{row}"], sheet3[f"B{row}"], sheet3[f"C{row}"] = \
            "Service", "Severity Level", "Count"

        for col in ("A", "B", "C"):
            c = sheet3[f"{col}{row}"]
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align

        row += 1
        start = row

        for (svc, sev), cnt in service_map.most_common(10):
            sheet3.cell(row, 1).value = svc
            sheet3.cell(row, 2).value = sev
            sheet3.cell(row, 3).value = cnt
            row += 1

        apply_border(sheet3, f"A{start-1}:C{row-1}")

        # ---------------------------------------------------------------------
        # RETURN EXCEL FILE
        # ---------------------------------------------------------------------
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "Problem_Analysis_Report.xlsx"
    


    # ===================================================================================
    # ✅ PDF FORMAT (FINAL VERSION – FIXED PIE LABELS + PAGE LAYOUT)
    # ===================================================================================
    if fmt == "pdf":
        buffer = io.BytesIO()

        pdf = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=20,
            rightMargin=20,
            topMargin=30,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        story = []

        # --------------------------------------------------------
        # ✅ TITLE BLOCK
        # --------------------------------------------------------
        story.append(Paragraph(f"<b><font size=18>Problem Analysis Report - {management_zone}</font></b>", styles["Title"]))
        story.append(Spacer(1, 12))
        story.append(Paragraph("Prepared by - KarvTech Team", styles["Heading4"]))
        story.append(Paragraph(f"Date - {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", styles["Heading4"]))
        story.append(Spacer(1, 18))

        # --------------------------------------------------------
        # ✅ TOP 20 PROBLEMS TABLE
        # --------------------------------------------------------
        story.append(Paragraph("<b>Top 20 Problems by Event</b>", styles["Heading3"]))

        top_table = [["Event Name", "Count"]]
        for t, c in title_counts.most_common(20):
            top_table.append([t, str(c)])

        T = Table(top_table, colWidths=[300, 80])
        T.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(T)
        story.append(Spacer(1, 20))

        # --------------------------------------------------------
        # ✅ PIE CHART (FIXED LABELS + FITS PAGE WIDTH)
        # --------------------------------------------------------
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.graphics.shapes import Drawing, String

        drawing = Drawing(520, 320)  # Increased canvas width for labels
        pie = Pie()

        pie.x = 150         # shift to the right to avoid left clipping
        pie.y = 50
        pie.width = 220
        pie.height = 220

        pie.data = [c for _, c in title_counts.most_common(20)]
        pie.labels = [t for t, _ in title_counts.most_common(20)]

        pie.sideLabels = True
        pie.simpleLabels = False
        pie.slices.strokeWidth = 0.5
        pie.slices.strokeColor = colors.black
        pie.slices.popout = 2
        pie.slices.fontName = "Helvetica"
        pie.slices.fontSize = 8
        pie.slices.fontColor = colors.black

        # Optional small title under chart
        drawing.add(String(180, 15, "Top 20 Problems by Event (Pie Chart)", fontSize=9))
        drawing.add(pie)

        story.append(drawing)
        story.append(Spacer(1, 20))

        # --------------------------------------------------------
        # ✅ PROBLEMATIC ENTITIES TABLE
        # --------------------------------------------------------
        story.append(Paragraph("<b>Problematic Entities</b>", styles["Heading3"]))

        ent_table = [["Entity Name", "Count", "Event Name"]]
        for name, cnt in entity_counter.most_common(40):
            evt = next((evt for en, evt in entity_event_map if en == name), "")
            ent_table.append([name, cnt, evt])

        T2 = Table(ent_table, repeatRows=1, colWidths=[200, 50, 200])
        T2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(T2)
        story.append(Spacer(1, 18))

        # --------------------------------------------------------
        # ✅ DETAILED PROBLEMS (NEW PAGE)
        # --------------------------------------------------------
        from reportlab.platypus import PageBreak
        story.append(PageBreak())
        story.append(Paragraph("<b>Detailed Problems</b>", styles["Heading3"]))

        detail = [
            ["Display ID", "Title", "Impact", "Entity Name", "Start IST", "End IST"]
        ]

        for row in all_rows:
            detail.append([
                row[0],  # Display ID
                row[1],  # Title
                row[2],  # Impact
                row[6],  # Entity Name
                row[7],  # Start IST
                row[8],  # End IST
            ])

        T4 = Table(
            detail,
            repeatRows=1,
            colWidths=[70, 140, 70, 150, 80, 80],
            splitByRow=True
        )

        T4.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(T4)

        # --------------------------------------------------------
        # ✅ BUILD PDF
        # --------------------------------------------------------
        pdf.build(story)
        buffer.seek(0)
        return buffer, "application/pdf", "Problem_Analysis_Report.pdf"


    raise ValueError(f"Unsupported format: {fmt}")


# ====================================================================================
# PROBLEM ANALYSIS PAGE (Validate → Download)
# ====================================================================================

@login_required(login_url='login')
def problem_data(request):

    tenant_url = (request.POST.get("tenant_url") or request.GET.get("tenant_url") or "").strip().rstrip("/")
    token = (request.POST.get("access_token") or request.GET.get("access_token") or "").strip()
    timeframe = request.POST.get("timeframe") or request.GET.get("timeframe") or "24h"
    action = request.POST.get("action")
    file_format = request.POST.get("format") or request.GET.get("format")

    # Normalize tenant
    if "apps.dynatrace.com" in tenant_url:
        tenant_url = tenant_url.replace("apps.dynatrace.com", "live.dynatrace.com")

    # Step 1 — Validate Tenant & Token
    if request.method == "POST" and action == "validate":
        try:
            zones = _fetch_management_zones(tenant_url, token)
        except Exception as e:
            return render(request, "problem_data.html", {
                "error": f"Validation failed: {str(e)}",
                "tenant_url": tenant_url,
                "access_token": token
            })

        return render(request, "problem_data.html", {
            "validated": True,
            "zones": zones,
            "tenant_url": tenant_url,
            "access_token": token,
            "timeframe": timeframe,
            "message": f"✔ Tenant validated — {len(zones)} zones found."
        })

    # Step 2 — Download Report
    if request.method == "POST" and action == "download":
        management_zone = request.POST.get("management_zone", "All")

        problems = _build_problem_dataset(tenant_url, token, timeframe, management_zone)
        if not problems:
            return render(request, "problem_data.html", {
                "validated": True,
                "tenant_url": tenant_url,
                "access_token": token,
                "message": "No problem data found for selected inputs."
            })

        # buf, mime, fname = _render_problem_report_bytes(problems, file_format)
        management_zone = request.POST.get("management_zone", "")
        buf, mime, fname = _render_problem_report_bytes(
            problems, 
            file_format,
            management_zone
        )

        response = HttpResponse(buf.getvalue(), content_type=mime)
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        return response


    return render(request, "problem_data.html")



# ====================================================================================
# FUNCTION USED BY APSCHEDULER (scheduler.py imports this)
# ====================================================================================
def generate_problem_analysis_report(
    tenant_url, access_token, management_zone, timeframe, report_format, email
):
    """
    Email contains 4 HTML tables:
      1. Top 10 Categories
      2. Top 10 Impacted Entities (with Severity)
      3. Top 10 Infrastructure Problems (impactLevel = INFRASTRUCTURE)
      4. Top 10 Services Problems (impactLevel = SERVICES)
    """
    try:
        # ------------------------------------------------------------
        # FETCH PROBLEM DATASET
        # ------------------------------------------------------------
        dataset = _build_problem_dataset(
            tenant_url, access_token, timeframe, management_zone
        )

        if not dataset:
            logger.info("⚠ No problems found — skipping email.")
            return

        # ------------------------------------------------------------
        # GENERATE ATTACHMENT
        # ------------------------------------------------------------
        buf, mime, fname = _render_problem_report_bytes(
            dataset, report_format, management_zone
        )

        # ------------------------------------------------------------
        # PREPARE SUMMARY DATA
        # ------------------------------------------------------------
        from collections import Counter, defaultdict

        # ========== TABLE 1: CATEGORIES ==========
        title_counts = Counter(
            [p.get("title", "") for p in dataset if p.get("title")]
        )

        # ========== TABLE 2: IMPACTED ENTITIES (ENTITY + SEVERITY) ==========
        entity_severity_counts = {}

        for p in dataset:
            if p.get("affectedEntities"):
                en = p["affectedEntities"][0].get("name", "")
                sev = p.get("severityLevel", "UNKNOWN")

                if en:
                    key = (en, sev)
                    entity_severity_counts[key] = entity_severity_counts.get(key, 0) + 1

        # ========== TABLE 3: INFRASTRUCTURE PROBLEMS ==========
        infra_map = {}

        for p in dataset:
            if p.get("impactLevel") == "INFRASTRUCTURE":
                name = p.get("affectedEntities", [{}])[0].get("name", "")
                sev = p.get("severityLevel", "UNKNOWN")
                if name:
                    key = (name, sev)
                    infra_map[key] = infra_map.get(key, 0) + 1

        # ========== TABLE 4: SERVICES PROBLEMS ==========
        service_map = {}

        for p in dataset:
            if p.get("impactLevel") == "SERVICES":
                name = p.get("affectedEntities", [{}])[0].get("name", "")
                sev = p.get("severityLevel", "UNKNOWN")
                if name:
                    key = (name, sev)
                    service_map[key] = service_map.get(key, 0) + 1

        # ------------------------------------------------------------
        # BUILD HTML BODY
        # ------------------------------------------------------------
        html_body = f"""
        <html>
        <body style="font-family: Arial; font-size: 14px; color:#333;">


        <p>Hello Team,<br>
        Please find below the summary of the <b>Problem Analysis Report</b>.
        The full detailed report is attached.</p>

        <p>
        <b>Management Zone:</b> {management_zone}<br>
        <b>Format:</b> {report_format.upper()}
        </p>

        <hr>

        <!-- ========================================================= -->
        <!--        TABLE 1 : Top 10 Categories                        -->
        <!-- ========================================================= -->
        <h3 style="color:#003366;">Top 10 Problems by Categories</h3>

        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:85%;">
        <tr style="background:#4F81BD; color:white;">
            <th>Category</th>
            <th>Count</th>
        </tr>
        """

        for title, cnt in title_counts.most_common(10):
            html_body += f"<tr><td>{title}</td><td>{cnt}</td></tr>"

        html_body += """
        </table><br><br>

        <!-- ========================================================= -->
        <!--        TABLE 2 : Top 10 Problems by Impacted Entities                 -->
        <!-- ========================================================= -->
        <h3 style="color:#003366;">Top 10 Problems by Impacted Entities</h3>

        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:85%;">
        <tr style="background:#4F81BD; color:white;">
            <th>Entity</th>
            <th>Severity Level</th>
            <th>Count</th>
        </tr>
        """

        for (entity, sev), cnt in sorted(entity_severity_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            html_body += f"<tr><td>{entity}</td><td>{sev}</td><td>{cnt}</td></tr>"

        html_body += """
        </table><br><br>

        <!-- ========================================================= -->
        <!--        TABLE 3 : Top 10 Infrastructure Problems           -->
        <!-- ========================================================= -->
        <h3 style="color:#003366;">Top 10 Infrastructure Problems</h3>

        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:85%;">
        <tr style="background:#4F81BD; color:white;">
            <th>Infrastructure</th>
            <th>Severity Level</th>
            <th>Count</th>
        </tr>
        """

        for (infra, sev), cnt in sorted(infra_map.items(), key=lambda x: x[1], reverse=True)[:10]:
            html_body += f"<tr><td>{infra}</td><td>{sev}</td><td>{cnt}</td></tr>"

        html_body += """
        </table><br><br>

        <!-- ========================================================= -->
        <!--        TABLE 4 : Top 10 Services Problems                 -->
        <!-- ========================================================= -->
        <h3 style="color:#003366;">Top 10 Services Problems</h3>

        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse; width:85%;">
        <tr style="background:#4F81BD; color:white;">
            <th>Services</th>
            <th>Severity Level</th>
            <th>Count</th>
        </tr>
        """

        for (svc, sev), cnt in sorted(service_map.items(), key=lambda x: x[1], reverse=True)[:10]:
            html_body += f"<tr><td>{svc}</td><td>{sev}</td><td>{cnt}</td></tr>"

        html_body += """
        </table>

        <br><br>
        Regards,<br>
        <b>KarvTech Automated Reporting System</b>

        </body>
        </html>
        """

        # ------------------------------------------------------------
        # SEND EMAIL
        # ------------------------------------------------------------
        message = EmailMessage(
            subject=f"Problem Analysis Report ({report_format.upper()})",
            body=html_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email]
        )

        message.content_subtype = "html"
        message.attach(fname, buf.getvalue(), mime)
        message.send(fail_silently=False)

        logger.info(f"✅ Report emailed to {email}")

    except Exception as e:
        logger.info(f"❌ Scheduler Report Error: {e}")


#######################################################################################################################################
#######################################################################################################################################
#######################################################################################################################################

# ====================================================================================
# EMAIL SCHEDULER UI (Validate Tenant → Create Schedule → List Schedules)
# ====================================================================================

IST = ZoneInfo("Asia/Kolkata")

def email_scheduler(request):
    context = {}

    # ------------------------------
    # VALIDATION STEP
    # ------------------------------
    if request.method == "POST" and request.POST.get("action") == "validate":
        report_type = request.POST.get("report_type", "").strip().lower()

        try:
            # USER MANAGEMENT
            if report_type == "user":
                account_uuid = request.POST.get("account_uuid", "").strip()
                client_id = request.POST.get("client_id", "").strip()
                client_secret = request.POST.get("client_secret", "").strip()

                if not all([account_uuid, client_id, client_secret]):
                    messages.error(request, "Please fill all required fields.")
                    return render(request, "email_scheduler.html")

                messages.success(request, "✅ User Management credentials validated successfully!")

                context.update({
                    "validated": True,
                    "report_type": "user",
                    "account_uuid": account_uuid,
                    "client_id": client_id,
                    "client_secret": client_secret,
                })

            # PROBLEM / SOFTWARE / CAPACITY
            else:
                tenant_url = _normalize_tenant(request.POST.get("tenant_url", ""))
                access_token = request.POST.get("access_token", "")

                zones = _fetch_management_zones(tenant_url, access_token)

                messages.success(request, f"✅ Tenant validated — {len(zones)} zones found.")

                context.update({
                    "validated": True,
                    "zones": zones,
                    "tenant_url": tenant_url,
                    "access_token": access_token,
                    "report_type": report_type,
                })

        except Exception as e:
            messages.error(request, f"❌ Validation failed: {e}")

    # ------------------------------
    # SCHEDULING STEP
    # ------------------------------
    elif request.method == "POST" and request.POST.get("action") == "schedule":
        try:
            next_run_str = request.POST.get("next_run")
            if not next_run_str:
                messages.error(request, "Next run date/time required.")
                context["schedules"] = ScheduledReport.objects.all().order_by("next_run")
                return render(request, "email_scheduler.html", context)

            aware_next_run = timezone.make_aware(datetime.fromisoformat(next_run_str), IST)

            ScheduledReport.objects.create(
                report_type=request.POST.get("report_type"),
                tenant_url=request.POST.get("tenant_url", ""),
                access_token=request.POST.get("access_token", ""),
                account_uuid=request.POST.get("account_uuid", ""),
                client_id=request.POST.get("client_id", ""),
                client_secret=request.POST.get("client_secret", ""),
                management_zone=request.POST.get("management_zone", ""),
                timeframe=request.POST.get("timeframe"),
                # recipient_email=request.POST.get("recipient_email"),
                recipient_email=",".join(
                    [e.strip() for e in request.POST.get("recipient_email","").split(",") if e.strip()]
                ),
                report_format=request.POST.get("report_format"),
                recurrence=request.POST.get("recurrence"),
                next_run=aware_next_run,
            )

            messages.success(request, "✅ Report scheduled successfully!")

            context["schedules"] = ScheduledReport.objects.all().order_by("next_run")
            return render(request, "email_scheduler.html", context)

        except Exception as e:
            messages.error(request, f"❌ Error creating schedule: {e}")
            context["schedules"] = ScheduledReport.objects.all().order_by("next_run")
            return render(request, "email_scheduler.html", context)

    # ------------------------------
    # INITIAL LOAD
    # ------------------------------
    context["schedules"] = ScheduledReport.objects.all().order_by("next_run")
    return render(request, "email_scheduler.html", context)



# ====================================================================================
# EDIT SCHEDULE
# ====================================================================================

def edit_schedule(request, pk):
    from django.utils import timezone
    from zoneinfo import ZoneInfo
    IST = ZoneInfo("Asia/Kolkata")

    schedule = get_object_or_404(ScheduledReport, pk=pk)
    zones = []  # Clean list of zone names only

    tenant_url = schedule.tenant_url
    access_token = schedule.access_token

    try:
        if tenant_url and access_token:
            raw_zones = _fetch_management_zones(tenant_url, access_token)
            # ✅ Extract only the "name" field from each zone dict
            zones = [z["name"] for z in raw_zones if isinstance(z, dict) and "name" in z]
    except Exception as e:
        logger.info(f"⚠️ Could not fetch management zones: {e}")

    if request.method == "POST":
        next_run_str = request.POST.get("next_run")
        if next_run_str:
            naive_dt = datetime.fromisoformat(next_run_str)
            schedule.next_run = timezone.make_aware(naive_dt, IST)

        schedule.management_zone = request.POST.get("management_zone")
        schedule.timeframe = request.POST.get("timeframe")
        schedule.report_format = request.POST.get("report_format")
        # schedule.recipient_email = request.POST.get("recipient_email")
        schedule.recipient_email = ",".join(
            [e.strip() for e in request.POST.get("recipient_email","").split(",") if e.strip()]
        )

        schedule.recurrence = request.POST.get("recurrence")

        schedule.save()
        messages.success(request, "✅ Schedule updated successfully!")
        return redirect("email_scheduler")

    return render(request, "edit_schedule.html", {
        "schedule": schedule,
        "zones": zones,
    })



# ====================================================================================
# DELETE SCHEDULE
# ====================================================================================

def delete_schedule(request, pk):
    job = get_object_or_404(ScheduledReport, pk=pk)
    job.delete()
    messages.success(request, "🗑️ Schedule deleted.")
    return redirect("email_scheduler")



############################################################################################################################################
############################################################################################################################################
############################################################################################################################################



# ===========================================
# DAILY ACTIVITY REPORT 
# ===========================================

def convert_to_ist(epoch_ms):
    """
    Convert epoch ms to IST string. Handles -1/None as Ongoing.
    Uses datetime.timezone (NOT django.utils.timezone) to avoid the 'module is not callable' bug.
    """
    if epoch_ms in (-1, None):
        return "Ongoing"
    try:
        IST = py_timezone(timedelta(hours=5, minutes=30))
        utc_time = datetime.fromtimestamp(epoch_ms / 1000, tz=py_timezone.utc)
        return utc_time.astimezone(IST).strftime("%d-%b-%Y %I:%M:%S %p")
    except Exception:
        return "Invalid Time"


def _normalize_timeframe(tf: str) -> str:
    tf = (tf or "").strip().lower()
    mapping = {
        "last 1 day": "24h", "last 7 days": "7d", "last 30 days": "30d",
        "1d": "24h", "24h": "24h", "7d": "7d", "30d": "30d",
        "today": "today", "yesterday": "yesterday", "1h": "1h"
    }
    return mapping.get(tf, "24h")


def _normalize_format(fmt: str | None) -> str:
    f = (fmt or "").strip().lower()
    mapping = {
        "csv": "csv",
        "excel": "excel", "xlsx": "excel", "xls": "excel",
        "pdf": "pdf",
        "pptx": "pptx", "ppt": "pptx", "powerpoint": "pptx"
    }
    return mapping.get(f, f)

def _get_timeframe_range(timeframe: str):
    """
    Returns (start_iso, end_iso) in UTC ISO8601.
    Uses datetime.timezone with alias py_timezone to avoid 'module is not callable'.
    """
    IST = py_timezone(timedelta(hours=5, minutes=30))
    now_utc = datetime.now(py_timezone.utc)
    now_ist = now_utc.astimezone(IST)

    tf = _normalize_timeframe(timeframe)

    if tf == "1h":
        start_utc, end_utc = now_utc - timedelta(hours=1), now_utc
    elif tf == "24h":
        start_utc, end_utc = now_utc - timedelta(hours=24), now_utc
    elif tf == "7d":
        start_utc, end_utc = now_utc - timedelta(days=7), now_utc
    elif tf == "30d":
        start_utc, end_utc = now_utc - timedelta(days=30), now_utc
    elif tf == "today":
        start_of_day_ist = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
        start_utc = start_of_day_ist.astimezone(py_timezone.utc)
        end_utc = now_utc
    elif tf == "yesterday":
        start_y_ist = (now_ist - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_y_ist = start_y_ist.replace(hour=23, minute=59, second=59, microsecond=0)
        start_utc = start_y_ist.astimezone(py_timezone.utc)
        end_utc = end_y_ist.astimezone(py_timezone.utc)
    else:
        start_utc, end_utc = now_utc - timedelta(hours=24), now_utc

    return (
        start_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
        end_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
    )


def _safe_filename(s: str) -> str:
    v = re.sub(r'[^A-Za-z0-9._-]+', '_', (s or "").strip())[:80]
    return v or datetime.now().strftime("%Y%m%d_%H%M%S")


# ===========================================
# Chart helpers
# ===========================================
def _timeseries_png(
    timestamps, avg_vals, max_vals, min_vals,
    title="CPU utilization split by hosts",
    y_label="CPU usage %", legend_base="CPU usage %",
    tick_interval="1h"):
    """
    Build a time-series PNG. Uses numpy & matplotlib. Headless-safe.
    """
    # Parse timestamps into datetimes
    x_dt = []
    for t in timestamps or []:
        if isinstance(t, datetime):
            x_dt.append(t)
            continue
        parsed = None
        if t is not None:
            s = str(t).strip()
            try:
                parsed = datetime.fromisoformat(s.replace("Z", "+00:00"))
            except Exception:
                pass
            if parsed is None:
                try:
                    parsed = datetime.strptime(s, "%Y-%m-%dT%H:%M:%S.%f")
                except Exception:
                    try:
                        parsed = datetime.strptime(s, "%Y-%m-%dT%H:%M:%S")
                    except Exception:
                        parsed = None
            if parsed is None:
                try:
                    v = float(s)
                    if v > 1e12:
                        parsed = datetime.fromtimestamp(v / 1000.0, tz=py_timezone.utc)
                    elif v > 1e9:
                        parsed = datetime.fromtimestamp(v, tz=py_timezone.utc)
                    else:
                        parsed = datetime.fromtimestamp(v, tz=py_timezone.utc)
                except Exception:
                    parsed = None
        x_dt.append(parsed)

    use_dates = any(isinstance(d, datetime) for d in x_dt)
    if use_dates:
        cleaned = []
        last = None
        for d in x_dt:
            if isinstance(d, datetime):
                last = d
                cleaned.append(d)
            else:
                cleaned.append(last or datetime.now(py_timezone.utc))
        x_nums = mdates.date2num(cleaned)
    else:
        x_nums = list(range(len(timestamps or [])))

    n = len(x_nums)

    def _to_floats(arr):
        out = []
        for i in range(n):
            try:
                out.append(float(arr[i]) if arr and i < len(arr) and arr[i] is not None else float("nan"))
            except Exception:
                out.append(float("nan"))
        return out

    y_avg = _to_floats(avg_vals)
    y_max = _to_floats(max_vals)
    y_min = _to_floats(min_vals)

    fig, ax = plt.subplots(figsize=(14, 4))
    ax.plot(x_nums, y_avg, linewidth=2.2, label=f'{legend_base} - Average')
    ax.plot(x_nums, y_max, linewidth=2.2, label=f'{legend_base} - Maximum')
    ax.plot(x_nums, y_min, linewidth=2.2, label=f'{legend_base} - Minimum')

    if use_dates:
        if tick_interval == "1h":
            locator = mdates.HourLocator(interval=1)
            fmt = "%H:%M"
        elif tick_interval == "30m":
            locator = mdates.MinuteLocator(byminute=range(0, 60, 30))
            fmt = "%H:%M"
        elif tick_interval == "15m":
            locator = mdates.MinuteLocator(byminute=range(0, 60, 15))
            fmt = "%H:%M"
        elif tick_interval == "5m":
            locator = mdates.MinuteLocator(byminute=range(0, 60, 5))
            fmt = "%H:%M"
        else:
            locator = mdates.AutoDateLocator()
            fmt = "%H:%M:%S"

        ax.xaxis.set_major_locator(locator)
        ax.xaxis.set_major_formatter(mdates.DateFormatter(fmt))
        fig.autofmt_xdate(rotation=35, ha='right')
    else:
        ax.set_xlim(min(x_nums) if x_nums else 0, max(x_nums) if x_nums else 1)

    ax.set_ylabel(y_label)
    ax.set_xlabel("")
    ax.set_title(title, fontsize=14, loc='left', pad=12)
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=100))
    ax.grid(axis='y', linestyle='--', alpha=0.35)
    ax.legend(loc='lower center', bbox_to_anchor=(0.5, -0.28), ncol=3, frameon=False, fontsize=9)
    plt.tight_layout(rect=[0, 0.02, 1, 0.98])

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _build_toplists(columns, table_rows):
    try:
        idx_name = columns.index("Host Name")
        idx_cpu = columns.index("CPU Usage (%)")
        idx_mem = columns.index("Memory Usage (%)")
        idx_disk = columns.index("Disk Usage (%)")
        idx_avl = columns.index("Host Availability (%)")
    except ValueError:
        return {"top_cpu": [], "top_mem": [], "top_disk": [], "worst_avl": []}

    def _safe_float(v):
        try:
            return float(v)
        except Exception:
            return None

    def _topN(idx_metric, reverse=True, n=10):
        pairs = []
        for r in table_rows:
            name = str(r[idx_name])
            val = _safe_float(r[idx_metric])
            if val is None:
                continue
            pairs.append((name, val))
        pairs.sort(key=lambda x: x[1], reverse=reverse)
        return pairs[:n]

    toplists = {
        "top_cpu": _topN(idx_cpu, True),
        "top_mem": _topN(idx_mem, True),
        "top_disk": _topN(idx_disk, True),
        "worst_avl": _topN(idx_avl, False),
    }
    return toplists


def _charts_from_table(columns, table_rows):
    try:
        idx_name = columns.index("Host Name")
    except ValueError:
        return {}

    idx_cpu = columns.index("CPU Usage (%)") if "CPU Usage (%)" in columns else None
    idx_mem = columns.index("Memory Usage (%)") if "Memory Usage (%)" in columns else None
    idx_disk = columns.index("Disk Usage (%)") if "Disk Usage (%)" in columns else None
    idx_avl = columns.index("Host Availability (%)") if "Host Availability (%)" in columns else None

    toplists = _build_toplists(columns, table_rows)
    top_cpu = toplists.get("top_cpu", [])
    top_mem = toplists.get("top_mem", [])
    top_disk = toplists.get("top_disk", [])

    def _barh_png(pairs, title):
        if not pairs:
            return None
        fig = plt.figure(figsize=(8, 4))
        labels = [p[0] for p in pairs][::-1]
        values = [p[1] for p in pairs][::-1]
        plt.barh(labels, values)
        plt.title(title)
        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=160, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    cpu_png = _barh_png(top_cpu, "Top 10 by CPU Usage (%)")
    mem_png = _barh_png(top_mem, "Top 10 by Memory Usage (%)")
    disk_png = _barh_png(top_disk, "Top 10 by Disk Usage (%)")

    # prepare per-host lines
    def _line_png(names, values, title, max_hosts=30, wrap_width=18):
        import textwrap as _tw
        if not names or not values:
            return None
        names = names[:max_hosts]
        values = values[:max_hosts]
        numeric_vals = []
        for v in values:
            try:
                numeric_vals.append(float(v))
            except Exception:
                numeric_vals.append(None)
        x = list(range(len(names)))
        wrapped_names = [_tw.fill(str(n), wrap_width) for n in names]

        fig_width = max(10, 0.35 * len(names) + 6)
        fig, ax = plt.subplots(figsize=(fig_width, 5))
        y_arr = np.array([np.nan if v is None else v for v in numeric_vals], dtype=float)
        ax.plot(x, y_arr, '-o', linewidth=2, markersize=6, markerfacecolor='white', markeredgecolor='black', zorder=3)
        ax.set_title(title, fontsize=13, fontweight='bold', pad=12)
        ax.set_xlabel("Hosts")
        ax.set_ylabel("Usage (%)")
        ax.set_xticks(x)
        ax.set_xticklabels(wrapped_names, rotation=45, ha='right', fontsize=8)
        ax.grid(axis='y', linestyle='--', alpha=0.5)

        vals_only = [v for v in numeric_vals if v is not None]
        if vals_only:
            if all(0 <= v <= 100 for v in vals_only):
                ax.set_ylim(0, 100)
            else:
                vmin, vmax = min(vals_only), max(vals_only)
                pad = max(1.0, (vmax - vmin) * 0.1)
                ax.set_ylim(vmin - pad, vmax + pad)

        plt.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=160, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    # sorted by CPU desc if present
    try:
        if idx_cpu is not None:
            sorted_rows = sorted(table_rows, key=lambda r: float(r[idx_cpu]) if r[idx_cpu] != "" else -1, reverse=True)
        else:
            sorted_rows = sorted(table_rows, key=lambda r: str(r[idx_name]).lower())
    except Exception:
        sorted_rows = table_rows

    host_names, cpu_vals, mem_vals, disk_vals, avl_vals = [], [], [], [], []
    for r in sorted_rows:
        name = str(r[idx_name])
        host_names.append(name)

        def _safe_val(idx):
            try:
                return float(r[idx]) if idx is not None and r[idx] != "" else 0.0
            except Exception:
                return 0.0

        cpu_vals.append(_safe_val(idx_cpu))
        mem_vals.append(_safe_val(idx_mem))
        disk_vals.append(_safe_val(idx_disk))
        avl_vals.append(_safe_val(idx_avl))

    cpu_line_png = _line_png(host_names, cpu_vals, "CPU Usage (Avg) per Host")
    mem_line_png = _line_png(host_names, mem_vals, "Memory Usage (Avg) per Host")
    disk_line_png = _line_png(host_names, disk_vals, "Disk Usage (Avg) per Host")
    avl_line_png = _line_png(host_names, avl_vals, "Availability (Avg) per Host")

    # attempt a small heatmap for top-CPU hosts (best effort)
    heatmap_png = None
    try:
        top_hosts = [n for n, _ in (top_cpu or [])][:15]
        heat_rows = []
        labels_rows = []
        for r in table_rows:
            name = str(r[idx_name])
            if name in top_hosts:
                try:
                    row_vals = [
                        float(r[idx_cpu]) if idx_cpu is not None and r[idx_cpu] != "" else 0.0,
                        float(r[idx_mem]) if idx_mem is not None and r[idx_mem] != "" else 0.0,
                        float(r[idx_disk]) if idx_disk is not None and r[idx_disk] != "" else 0.0,
                        float(r[idx_avl]) if idx_avl is not None and r[idx_avl] != "" else 0.0,
                    ]
                except Exception:
                    continue
                labels_rows.append(name)
                heat_rows.append(row_vals)

        if heat_rows:
            arr = np.array(heat_rows, dtype=float)
            fig = plt.figure(figsize=(6, 0.4 * len(labels_rows) + 1.5))
            plt.imshow(arr, aspect="auto")
            plt.xticks([0, 1, 2, 3], ["CPU%", "MEM%", "DISK%", "AVL%"])
            plt.yticks(range(len(labels_rows)), labels_rows)
            plt.title("Heatmap (Top CPU Hosts)")
            plt.tight_layout()
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=160, bbox_inches='tight')
            plt.close(fig)
            buf.seek(0)
            heatmap_png = buf.getvalue()
    except Exception:
        heatmap_png = None

    return {
        "cpu_png": cpu_png, "mem_png": mem_png, "disk_png": disk_png,
        "heatmap_png": heatmap_png,
        "cpu_line_png": cpu_line_png, "mem_line_png": mem_line_png,
        "disk_line_png": disk_line_png, "avl_line_png": avl_line_png,
        "_toplists": toplists
    }

# =====================================================================
#  EMAIL VERSION OF CAPACITY MANAGEMENT REPORT
#  (Excel + PDF support — SAME layout as manual download)
# =====================================================================
from django.conf import settings

def generate_capacity_management_report(
    tenant_url,
    access_token,
    management_zone,
    timeframe,
    report_format,
    email,
):
    """
    Generates the Capacity Management report (same as daily_activity
    manual download) and emails it as Excel or PDF.
    """
    import io
    import base64
    import re
    from datetime import date
    import requests
    from django.core.mail import EmailMessage

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, Image as RLImage
    )
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    # -----------------------------
    # 1. Time window & entity selector
    # -----------------------------
    start_iso, end_iso = _get_timeframe_range(timeframe)

    entity_selector = None
    if management_zone and management_zone.lower() != "all":
        entity_selector = _build_entity_selector_from_mz(management_zone)

    # -----------------------------
    # 2. Per-host aggregated metrics (same as daily_activity)
    # -----------------------------
    per_host = {}
    for metric_id, display, agg, _needs_ts in METRICS:
        rows = _query_metric_split_by_host(
            tenant_url, access_token, metric_id, agg,
            start_iso, end_iso, needs_timeslices=True,
            entity_selector=entity_selector,
        )
        base = display.replace("(%)", "").strip()
        min_col = f"{base} (Min %)" if "(%)" in display else f"{base} (Min)"
        max_col = f"{base} (Max %)" if "(%)" in display else f"{base} (Max)"

        for r in rows:
            rec = per_host.setdefault(
                r["host_id"],
                {"Host ID": r["host_id"], "Host Name": r["host_name"]},
            )
            rec[display] = round(r["value_avg"], 2)
            rec[min_col] = round(r["value_min"], 2)
            rec[max_col] = round(r["value_max"], 2)

    if not per_host:
        raise Exception("No metric data found for the selected inputs.")

    # -----------------------------
    # 3. Columns + table_rows (same as daily_activity)
    # -----------------------------
    base_columns = ["Host ID", "Host Name"]
    avg_columns = [m[1] for m in METRICS]
    extra_columns = []

    sample_key = next(iter(per_host))
    sample = per_host[sample_key]
    for _, display, *_ in METRICS:
        base = display.replace("(%)", "").strip()
        min_col = f"{base} (Min %)" if "(%)" in display else f"{base} (Min)"
        max_col = f"{base} (Max %)" if "(%)" in display else f"{base} (Max)"
        if min_col in sample:
            extra_columns.append(min_col)
        if max_col in sample:
            extra_columns.append(max_col)

    columns = base_columns + avg_columns + extra_columns

    table_rows = []
    for _, row in sorted(
        per_host.items(),
        key=lambda kv: (kv[1].get("Host Name") or kv[1]["Host ID"] or "").lower()
    ):
        clean = []
        for col in columns:
            val = row.get(col, "")
            if isinstance(val, float):
                val = round(val, 2)
            clean.append(val)
        table_rows.append(clean)

    # -----------------------------
    # 4. Charts (same as daily_activity)
    # -----------------------------
    charts = _charts_from_table(columns, table_rows)

    cpu_ts_png = mem_ts_png = disk_ts_png = None
    try:
        # CPU
        ts_timestamps, cpu_avg = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.cpu.usage", "avg",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        _, cpu_max = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.cpu.usage", "max",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        _, cpu_min = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.cpu.usage", "min",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        if ts_timestamps and (cpu_avg or cpu_max or cpu_min):
            min_len = min(len(ts_timestamps), len(cpu_avg), len(cpu_max), len(cpu_min))
            cpu_ts_png = _timeseries_png(
                ts_timestamps[:min_len], cpu_avg[:min_len], cpu_max[:min_len], cpu_min[:min_len],
                title="CPU utilization split by hosts", y_label="CPU usage %", legend_base="CPU usage %", tick_interval="1h"
            )
        else:
            cpu_ts_png = charts.get("cpu_line_png")

        # Memory
        mem_ts_timestamps, mem_avg = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.mem.usage", "avg",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        _, mem_max = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.mem.usage", "max",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        _, mem_min = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.mem.usage", "min",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        if mem_ts_timestamps and (mem_avg or mem_max or mem_min):
            min_len = min(len(mem_ts_timestamps), len(mem_avg), len(mem_max), len(mem_min))
            mem_ts_png = _timeseries_png(
                mem_ts_timestamps[:min_len], mem_avg[:min_len], mem_max[:min_len], mem_min[:min_len],
                title="Memory usage split by hosts", y_label="Memory usage %", legend_base="Memory usage %", tick_interval="1h"
            )
        else:
            mem_ts_png = charts.get("mem_line_png")

        # Disk
        disk_ts_timestamps, disk_avg = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.disk.usedPct", "avg",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        _, disk_max = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.disk.usedPct", "max",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        _, disk_min = _query_metric_timeseries(
            tenant_url, access_token, "builtin:host.disk.usedPct", "min",
            start_iso, end_iso, resolution="1m", entity_selector=entity_selector
        )
        if disk_ts_timestamps and (disk_avg or disk_max or disk_min):
            min_len = min(len(disk_ts_timestamps), len(disk_avg), len(disk_max), len(disk_min))
            disk_ts_png = _timeseries_png(
                disk_ts_timestamps[:min_len], disk_avg[:min_len], disk_max[:min_len], disk_min[:min_len],
                title="Disk usage split by hosts", y_label="Disk usage %", legend_base="Disk usage %", tick_interval="1h"
            )
        else:
            disk_ts_png = charts.get("disk_line_png")
    except Exception:
        cpu_ts_png = cpu_ts_png or charts.get("cpu_line_png")
        mem_ts_png = mem_ts_png or charts.get("mem_line_png")
        disk_ts_png = disk_ts_png or charts.get("disk_line_png")

    mz_display = management_zone or "All"

    # ------------------------------------------
    # 5. BUILD EXCEL (same as daily_activity Excel path)
    # ------------------------------------------
    if report_format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Host Metrics"

        # ----- Title Section -----
        title = f"Capacity Management Report For {mz_display if mz_display else 'All'}"
        ws.merge_cells("A1:H4")
        ws["A1"].value = title
        ws["A1"].font = Font(size=18, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

        ws.merge_cells("A5:H5")
        ws["A5"].value = "Prepared by - KarvTech Team"
        ws["A5"].font = Font(size=12, italic=True, bold=True)
        ws["A5"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

        ws.merge_cells("A6:H6")
        ws["A6"].value = f"Date - {date.today().strftime('%d-%m-%Y')}"
        ws["A6"].font = Font(size=12, italic=True, bold=True)
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

        # ----- Table Header -----
        start_row = 8
        headers = [
            "Host Name",
            "Host Availability %",
            "CPU Usage Max %",
            "CPU Usage Min %",
            "CPU Usage Avg %",
            "Memory Usage Max %",
            "Memory Usage Min %",
            "Memory Usage Avg %",
            "Disk Usage Max %",
            "Disk Usage Min %",
            "Disk Usage Avg %",
        ]
        thin = Side(border_style="thin", color="000000")

        for idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=idx, value=col_name)
            cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        def _fmt(v):
            if v is None or v == "":
                return ""
            try:
                return round(float(v), 2)
            except Exception:
                return v

        hosts_sorted = sorted(per_host.items(), key=lambda kv: (kv[1].get("Host Name") or "").lower())

        cpu_avg_key = "CPU Usage (%)"
        cpu_min_key = "CPU Usage (Min %)"
        cpu_max_key = "CPU Usage (Max %)"
        mem_avg_key = "Memory Usage (%)"
        mem_min_key = "Memory Usage (Min %)"
        mem_max_key = "Memory Usage (Max %)"
        disk_avg_key = "Disk Usage (%)"
        disk_min_key = "Disk Usage (Min %)"
        disk_max_key = "Disk Usage (Max %)"""

        # ----- Write Table Rows -----
        for _, rec in hosts_sorted:
            row_vals = [
                rec.get("Host Name", ""),
                _fmt(rec.get("Host Availability (%)", "")) / 100 if _fmt(rec.get("Host Availability (%)", "")) != "" else "",
                _fmt(rec.get(cpu_max_key, "")) / 100 if _fmt(rec.get(cpu_max_key, "")) != "" else "",
                _fmt(rec.get(cpu_min_key, "")) / 100 if _fmt(rec.get(cpu_min_key, "")) != "" else "",
                _fmt(rec.get(cpu_avg_key, "")) / 100 if _fmt(rec.get(cpu_avg_key, "")) != "" else "",
                _fmt(rec.get(mem_max_key, "")) / 100 if _fmt(rec.get(mem_max_key, "")) != "" else "",
                _fmt(rec.get(mem_min_key, "")) / 100 if _fmt(rec.get(mem_min_key, "")) != "" else "",
                _fmt(rec.get(mem_avg_key, "")) / 100 if _fmt(rec.get(mem_avg_key, "")) != "" else "",
                _fmt(rec.get(disk_max_key, "")) / 100 if _fmt(rec.get(disk_max_key, "")) != "" else "",
                _fmt(rec.get(disk_min_key, "")) / 100 if _fmt(rec.get(disk_min_key, "")) != "" else "",
                _fmt(rec.get(disk_avg_key, "")) / 100 if _fmt(rec.get(disk_avg_key, "")) != "" else "",
            ]

            ws.append(row_vals)
            r = ws.max_row

            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            for c in range(2, 12):
                ws.cell(row=r, column=c).number_format = "0.00%"

        # Autofit
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        # ----- Embed charts (same as daily_activity) -----
        def _embed(sheet, png_bytes, cell, width=950):
            try:
                pil = PILImage.open(io.BytesIO(png_bytes))
                if pil.mode not in ("RGB", "RGBA"):
                    pil = pil.convert("RGBA")
                buff = io.BytesIO()
                pil.save(buff, format="PNG")
                buff.seek(0)
                img = XLImage(buff)
                img.width = width
                img.height = int(width * (pil.height / pil.width))
                sheet.add_image(img, cell)
                rows_needed = int(img.height / 20) + 3
                return rows_needed
            except Exception:
                return 30

        chart_row = 14
        if cpu_ts_png:
            used = _embed(ws, cpu_ts_png, f"A{chart_row}")
            chart_row += used
        if mem_ts_png:
            used = _embed(ws, mem_ts_png, f"A{chart_row}")
            chart_row += used
        if disk_ts_png:
            used = _embed(ws, disk_ts_png, f"A{chart_row}")
            chart_row += used

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = "Capacity_Management.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        file_bytes = buf.getvalue()

    # ------------------------------------------
    # 6. BUILD PDF (same as daily_activity PDF path)
    # ------------------------------------------
    elif report_format == "pdf":
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title_style', parent=styles['Title'],
                                     fontSize=22, leading=24, alignment=0, spaceAfter=6)
        meta_style = ParagraphStyle('meta_style', parent=styles['Normal'],
                                    fontSize=9, leading=11, alignment=0, italic=True)
        small_wrap = ParagraphStyle('small_wrap', parent=styles['Normal'],
                                    fontSize=7, leading=8, spaceBefore=0, spaceAfter=0)
        header_wrap = ParagraphStyle('header_wrap', parent=styles['Normal'],
                                     fontSize=8, leading=9, alignment=1)

        elements = []
        report_title = f"Capacity Management Report For {mz_display if mz_display else 'All'}"
        elements.append(
            Table(
                [[Paragraph(f"<b>{report_title}</b>", title_style)],
                 [Paragraph("Prepared by - KarvTech Team", meta_style)],
                 [Paragraph(f"Date - {date.today().strftime('%d-%m-%Y')}", meta_style)]],
                colWidths=[doc.width]
            )
        )
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"Timeframe: {timeframe} (UTC: {start_iso} → {end_iso})", styles['Normal']))
        elements.append(Spacer(1, 8))

        # Remove Host ID
        host_id_index = None
        for i, c in enumerate(columns):
            if re.search(r'host\s*id', str(c), re.I):
                host_id_index = i
                break
        if host_id_index is not None:
            filtered_columns = [c for j, c in enumerate(columns) if j != host_id_index]
            filtered_rows = [[val for j, val in enumerate(row) if j != host_id_index] for row in table_rows]
        else:
            filtered_columns = columns[:]
            filtered_rows = [row[:] for row in table_rows]

        # Remove Host Availability Min/Max
        remove_patterns = re.compile(r'host\s*availability.*\b(min|max)\b', re.I)
        keep_index_pairs = [(i, c) for i, c in enumerate(filtered_columns) if not remove_patterns.search(str(c))]
        keep_indexes = [i for i, _ in keep_index_pairs]
        pdf_columns = [filtered_columns[i] for i in keep_indexes]
        pdf_table_rows = [[row[i] if i < len(row) else '' for i in keep_indexes] for row in filtered_rows]

        # Reorder columns
        indexed = list(enumerate(pdf_columns))
        ordered_indexes, ordered_columns = [], []

        # Host Name
        for i, h in indexed:
            if re.search(r'host\s*name', str(h), re.I):
                ordered_indexes.append(i)
                ordered_columns.append(h)
                break

        # Host Availability (overall)
        for i, h in indexed:
            if i in ordered_indexes:
                continue
            if re.search(r'host\s*availability(?!.*\b(min|max)\b)', str(h), re.I):
                ordered_indexes.append(i)
                ordered_columns.append(h)
                break

        metrics = [
            ('CPU', r'cpu'),
            ('Memory', r'\b(mem|memory)\b'),
            ('Disk', r'disk'),
        ]
        for _, metric_re in metrics:
            # Usage
            for i, h in indexed:
                if i in ordered_indexes:
                    continue
                if re.search(metric_re, str(h), re.I) and re.search(r'usage', str(h), re.I) and not re.search(r'\b(min|max)\b', str(h), re.I):
                    ordered_indexes.append(i)
                    ordered_columns.append(h)
                    break
            # Min
            for i, h in indexed:
                if i in ordered_indexes:
                    continue
                if re.search(metric_re, str(h), re.I) and re.search(r'\bmin\b', str(h), re.I):
                    ordered_indexes.append(i)
                    ordered_columns.append(h)
                    break
            # Max
            for i, h in indexed:
                if i in ordered_indexes:
                    continue
                if re.search(metric_re, str(h), re.I) and re.search(r'\bmax\b', str(h), re.I):
                    ordered_indexes.append(i)
                    ordered_columns.append(h)
                    break

        for i, h in indexed:
            if i not in ordered_indexes:
                ordered_indexes.append(i)
                ordered_columns.append(h)

        final_rows = [[row[i] if i < len(row) else '' for i in ordered_indexes] for row in pdf_table_rows]
        data = [ordered_columns] + final_rows

        # Convert to Paragraphs
        for r in range(len(data)):
            for c in range(len(data[r])):
                text = "" if data[r][c] is None else str(data[r][c])
                if r == 0:
                    data[r][c] = Paragraph(f"<b>{text}</b>", header_wrap)
                else:
                    data[r][c] = Paragraph(text, small_wrap)

        num_cols = len(ordered_columns) if ordered_columns else 1
        rel_weights = [3.0] + [1.0] * (num_cols - 1)
        total_weight = sum(rel_weights)
        available_width = doc.width
        col_widths = [available_width * (w / total_weight) for w in rel_weights]

        min_first = 140
        if num_cols >= 1 and col_widths[0] < min_first:
            remaining = available_width - min_first
            if remaining < 0:
                remaining = 30 * (num_cols - 1)
                min_first = available_width - remaining
            col_widths[0] = min_first
            for i in range(1, num_cols):
                col_widths[i] = max(30, remaining * (rel_weights[i] / (total_weight - rel_weights[0])))

        numeric_regex = re.compile(r'(cpu|mem|memory|disk|usage|min|max|%)', re.I)
        numeric_cols = []
        for i, h in enumerate(ordered_columns):
            if numeric_regex.search(str(h)):
                numeric_cols.append(i)

        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
        tbl_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor("#666666")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ])

        if num_cols >= 1:
            tbl_style.add('ALIGN', (0, 1), (0, -1), 'LEFT')
        for idx in numeric_cols:
            tbl_style.add('ALIGN', (idx, 1), (idx, -1), 'RIGHT')
            tbl_style.add('ALIGN', (idx, 0), (idx, 0), 'CENTER')

        table.setStyle(tbl_style)
        elements.append(table)
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("<b>Visualizations</b>", styles['Heading2']))
        elements.append(Spacer(1, 6))

        # charts
        if cpu_ts_png:
            elements.append(Paragraph("CPU utilization split by hosts", styles['Heading3']))
            try:
                pil_img = PILImage.open(io.BytesIO(cpu_ts_png))
                orig_w, orig_h = pil_img.size
                target_width = doc.width
                ratio = (orig_h / float(orig_w)) if orig_w else 0.4
                target_height = target_width * ratio
                elements.append(RLImage(io.BytesIO(cpu_ts_png), width=target_width, height=target_height))
            except Exception:
                elements.append(RLImage(io.BytesIO(cpu_ts_png), width=doc.width, height=3.4 * inch))
            elements.append(Spacer(1, 10))

        if mem_ts_png:
            elements.append(Paragraph("Memory utilization split by hosts", styles['Heading3']))
            try:
                pil_img = PILImage.open(io.BytesIO(mem_ts_png))
                orig_w, orig_h = pil_img.size
                target_width = doc.width
                ratio = (orig_h / float(orig_w)) if orig_w else 0.4
                target_height = target_width * ratio
                elements.append(RLImage(io.BytesIO(mem_ts_png), width=target_width, height=target_height))
            except Exception:
                elements.append(RLImage(io.BytesIO(mem_ts_png), width=doc.width, height=3.4 * inch))
            elements.append(Spacer(1, 10))

        if disk_ts_png:
            elements.append(Paragraph("Disk utilization split by hosts", styles['Heading3']))
            try:
                pil_img = PILImage.open(io.BytesIO(disk_ts_png))
                orig_w, orig_h = pil_img.size
                target_width = doc.width
                ratio = (orig_h / float(orig_w)) if orig_w else 0.4
                target_height = target_width * ratio
                elements.append(RLImage(io.BytesIO(disk_ts_png), width=target_width, height=target_height))
            except Exception:
                elements.append(RLImage(io.BytesIO(disk_ts_png), width=doc.width, height=3.4 * inch))
            elements.append(Spacer(1, 10))

        doc.build(elements)
        buf.seek(0)
        fname = "Capacity_Management.pdf"
        mime = "application/pdf"
        file_bytes = buf.getvalue()

    else:
        raise Exception(f"Unsupported format '{report_format}'")

    # ------------------------------------------
    # 7. SEND EMAIL WITH ATTACHMENT
    # ------------------------------------------
    subject = f"Capacity Management Report - {mz_display}"
    body = (
        "Hello,\n\n"
        "Attached is your scheduled Capacity Management report.\n\n"
        f"Timeframe: {timeframe}\n"
        f"Management Zone: {mz_display}\n\n"
        "Regards,\nKarvTech Automated Reporting"
    )

    msg = EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, [email])
    msg.attach(fname, file_bytes, mime)
    msg.send(fail_silently=False)

# ===========================================
# Dynatrace metrics wiring + API wrappers
# ===========================================
METRICS = [
    ("builtin:host.availability", "Host Availability (%)", "avg", True),
    ("builtin:host.cpu.usage",    "CPU Usage (%)",         "avg", True),
    ("builtin:host.mem.usage",    "Memory Usage (%)",      "avg", True),
    ("builtin:host.disk.usedPct", "Disk Usage (%)",        "avg", True),
]


def _fetch_management_zones(tenant_url: str, token: str):
    url = f"{tenant_url}/api/config/v1/managementZones"
    headers = {"Authorization": f"Api-Token {token}", "Accept": "application/json"}
    resp = requests.get(url, headers=headers, timeout=60)
    resp.raise_for_status()
    payload = resp.json() or {}
    values = payload.get("values") or []
    zones = [{"id": v.get("id"), "name": v.get("name")} for v in values if v.get("name")]
    zones.sort(key=lambda z: z["name"].lower())
    return zones


def _build_entity_selector_from_mz(mz_value: str | None) -> str | None:
    if not mz_value or mz_value.strip().lower() == "all":
        return None
    v = mz_value.strip()
    if re.fullmatch(r"-?\d+", v):
        return f'type("HOST"),mzId("{v}")'
    if re.fullmatch(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", v):
        return f'type("HOST"),mzId("{v}")'
    safe_name = v.replace('"', '\\"')
    return f'type("HOST"),mzName("{safe_name}")'

def get_real_host_name(tenant_url, token, host_id):
    """
    Fetch real hostname for each host_id using /api/v2/entities endpoint.
    Returns displayName or name or fallback to host_id.
    """
    url = f"{tenant_url}/api/v2/entities/{host_id}"
    headers = {"Authorization": f"Api-Token {token}"}

    try:
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code == 200:
            data = r.json() or {}
            return (
                data.get("displayName")
                or data.get("properties", {}).get("hostname")
                or data.get("properties", {}).get("detectedName")
                or data.get("entityId")
                or host_id
            )
    except:
        pass

    return host_id

def _query_metric_split_by_host(base_url, token, metric_id, agg,
                                start_iso, end_iso, needs_timeslices=True,
                                entity_selector: str | None = None):

    url = f"{base_url}/api/v2/metrics/query"
    headers = {"Authorization": f"Api-Token {token}", "Accept": "application/json"}
    selector = f'{metric_id}:{agg}:splitBy("dt.entity.host")'
    params = {
        "metricSelector": selector,
        "from": start_iso,
        "to": end_iso,
        "resolution": "5m",
        "fields": "entities",
        "pageSize": 4000
    }
    if entity_selector:
        params["entitySelector"] = entity_selector

    resp = requests.get(url, headers=headers, params=params, timeout=60)
    resp.raise_for_status()
    data = resp.json()

    PCT_METRICS = {
        "builtin:host.availability",
        "builtin:host.cpu.usage",
        "builtin:host.mem.usage",
        "builtin:host.disk.usedPct",
    }

    def _norm_pct(val):
        v = float(val)
        if metric_id in PCT_METRICS and v <= 1.0:
            v *= 100.0
        return max(0.0, min(100.0, v)) if metric_id in PCT_METRICS else v

    out = []

    for res in data.get("result", []):
        metric_entities = {e.get("id"): e.get("name") for e in res.get("entities", [])}

        for series in res.get("data", []):
            host_id = (series.get("dimensions") or [""])[0]

            # -------- REAL FIX --------
            # Fetch real displayName
            real_name = get_real_host_name(base_url, token, host_id)
            metric_name = (metric_entities.get(host_id) or "").strip()

            host_name = real_name or metric_name or host_id
            # --------------------------

            vals = []
            for raw in (series.get("values") or []):
                if raw is not None:
                    vals.append(_norm_pct(raw))

            if not vals:
                continue

            out.append({
                "host_id": host_id,
                "host_name": host_name,
                "value_min": min(vals),
                "value_max": max(vals),
                "value_avg": sum(vals) / len(vals)
            })

    return out




def _query_metric_timeseries(base_url, token, metric_id, agg, start_iso, end_iso,
                             resolution="5m", entity_selector: str | None = None):
    url = f"{base_url}/api/v2/metrics/query"
    selector = f"{metric_id}:{agg}"
    params = {
        "metricSelector": selector,
        "from": start_iso, "to": end_iso,
        "resolution": resolution, "pageSize": 4000
    }
    if entity_selector:
        params["entitySelector"] = entity_selector

    headers = {"Authorization": f"Api-Token {token}", "Accept": "application/json"}
    resp = requests.get(url, headers=headers, params=params, timeout=60)
    resp.raise_for_status()
    data = resp.json() or {}

    timestamps, values = [], []
    try:
        for res in data.get("result", []):
            for series in res.get("data", []):
                ts = series.get("timestamps") or []
                vals = series.get("values") or []
                if ts and vals:
                    timestamps = ts
                    values = vals
                    break
            if timestamps:
                break
    except Exception:
        pass

    PCT_METRICS = {
        "builtin:host.availability",
        "builtin:host.cpu.usage",
        "builtin:host.mem.usage",
        "builtin:host.disk.usedPct",
    }

    def _norm(v):
        try:
            vv = float(v)
            if metric_id in PCT_METRICS and vv <= 1.0:
                vv *= 100.0
            if metric_id in PCT_METRICS:
                vv = max(0.0, min(100.0, vv))
            return vv
        except Exception:
            return None

    norm_vals = [_norm(v) for v in values]
    return timestamps, norm_vals


# ===========================================
# Main view
# ===========================================
@login_required(login_url='login')
def daily_activity(request):
    tenant_url = (request.POST.get("tenant_url") or request.GET.get("tenant_url") or request.session.get("dt_tenant_url") or "").strip().rstrip("/")
    access_token = (request.POST.get("access_token") or request.GET.get("access_token") or request.session.get("dt_access_token") or "").strip()
    timeframe = (request.POST.get("timeframe") or request.GET.get("timeframe") or request.session.get("dt_timeframe") or "Last 1 Day")
    action = request.POST.get("action")
    file_format = _normalize_format(request.POST.get("format") or request.GET.get("format") or "csv")

    if tenant_url and "apps.dynatrace.com" in tenant_url:
        tenant_url = tenant_url.replace("apps.dynatrace.com", "live.dynatrace.com")

    # Validate
    if request.method == "POST" and action == "validate":
        if not tenant_url or not access_token:
            return render(request, "daily_activity.html", {
                "error": "Please enter Tenant URL and Access Token, then click Validate.",
                "tenant_url": tenant_url, "timeframe": timeframe
            })
        try:
            zones = _fetch_management_zones(tenant_url, access_token)
        except requests.HTTPError as e:
            return render(request, "daily_activity.html", {
                "error": f"Validation failed: {e.response.status_code} - {e.response.text[:300]}",
                "tenant_url": tenant_url, "timeframe": timeframe
            })
        except Exception as e:
            return render(request, "daily_activity.html", {
                "error": f"Validation failed: {str(e)}",
                "tenant_url": tenant_url, "timeframe": timeframe
            })

        request.session["dt_tenant_url"] = tenant_url
        request.session["dt_access_token"] = access_token
        request.session["dt_timeframe"] = timeframe

        return render(request, "daily_activity.html", {
            "validated": True, "zones": zones,
            "tenant_url": tenant_url, "timeframe": timeframe
        })

    # Download / Preview
    if request.method == "POST" and action in ("download", "preview"):
        if not tenant_url:
            tenant_url = request.session.get("dt_tenant_url", "")
        if not access_token:
            access_token = request.session.get("dt_access_token", "")
        if not timeframe:
            timeframe = request.session.get("dt_timeframe", "Last 1 Day")

        if not tenant_url or not access_token:
            return render(request, "daily_activity.html", {
                "error": "Tenant URL and Access Token are required.",
                "tenant_url": tenant_url, "timeframe": timeframe
            })

        management_zone = (request.POST.get("management_zone") or "").strip()
        management_zone_name = (request.POST.get("management_zone_name") or "").strip()
        request.session["dt_management_zone"] = management_zone

        start_iso, end_iso = _get_timeframe_range(timeframe)

        entity_selector = None
        if management_zone and management_zone.lower() != "all":
            entity_selector = _build_entity_selector_from_mz(management_zone)

        # per-host summary
        per_host = {}
        try:
            for metric_id, display, agg, _needs_ts in METRICS:
                rows = _query_metric_split_by_host(
                    tenant_url, access_token, metric_id, agg, start_iso, end_iso,
                    needs_timeslices=True, entity_selector=entity_selector
                )
                base = display.replace("(%)", "").strip()
                min_col = f"{base} (Min %)" if "(%)" in display else f"{base} (Min)"
                max_col = f"{base} (Max %)" if "(%)" in display else f"{base} (Max)"

                for r in rows:
                    rec = per_host.setdefault(r["host_id"], {"Host ID": r["host_id"], "Host Name": r["host_name"]})
                    rec[display] = round(r["value_avg"], 2)
                    rec[min_col] = round(r["value_min"], 2)
                    rec[max_col] = round(r["value_max"], 2)
        except requests.HTTPError as e:
            return render(request, "daily_activity.html", {
                "error": f"Dynatrace API error: {e.response.status_code} - {e.response.text[:300]}",
                "validated": True, "tenant_url": tenant_url, "timeframe": timeframe
            })
        except Exception as e:
            return render(request, "daily_activity.html", {
                "error": f"Unexpected error: {str(e)}",
                "validated": True, "tenant_url": tenant_url, "timeframe": timeframe
            })

        if not per_host:
            return render(request, "daily_activity.html", {
                "message": "No metric data found for the selected inputs.",
                "validated": True, "tenant_url": tenant_url, "timeframe": timeframe
            })

        base_columns = ["Host ID", "Host Name"]
        avg_columns = [m[1] for m in METRICS]
        extra_columns = []

        sample_key = next(iter(per_host))
        sample = per_host[sample_key]
        for _, display, *_ in METRICS:
            base = display.replace("(%)", "").strip()
            min_col = f"{base} (Min %)" if "(%)" in display else f"{base} (Min)"
            max_col = f"{base} (Max %)" if "(%)" in display else f"{base} (Max)"
            if min_col in sample:
                extra_columns.append(min_col)
            if max_col in sample:
                extra_columns.append(max_col)

        columns = base_columns + avg_columns + extra_columns

        table_rows = []
        for _, row in sorted(per_host.items(), key=lambda kv: (kv[1].get("Host Name") or kv[1]["Host ID"] or "").lower()):
            clean = []
            for col in columns:
                val = row.get(col, "")
                if isinstance(val, float):
                    val = round(val, 2)
                clean.append(val)
            table_rows.append(clean)

        label_mz = _safe_filename(management_zone or "All")
        label_tf = _safe_filename(timeframe)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')

        # CSV download
        if file_format == "csv" and action == "download":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            writer.writerows(table_rows)
            resp = HttpResponse(output.getvalue(), content_type="text/csv")
            resp["Content-Disposition"] = f'attachment; filename="host_metrics_{label_mz}_{label_tf}_{timestamp}.csv"'
            return resp

        # charts from table
        charts = _charts_from_table(columns, table_rows)

        # timeseries summary across timeframe
        cpu_ts_png = mem_ts_png = disk_ts_png = None
        try:
            # CPU
            ts_timestamps, cpu_avg = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.cpu.usage", "avg",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            _, cpu_max = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.cpu.usage", "max",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            _, cpu_min = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.cpu.usage", "min",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            if ts_timestamps and (cpu_avg or cpu_max or cpu_min):
                min_len = min(len(ts_timestamps), len(cpu_avg), len(cpu_max), len(cpu_min))
                cpu_ts_png = _timeseries_png(
                    ts_timestamps[:min_len], cpu_avg[:min_len], cpu_max[:min_len], cpu_min[:min_len],
                    title="CPU utilization split by hosts", y_label="CPU usage %", legend_base="CPU usage %", tick_interval="1h"
                )
            else:
                cpu_ts_png = charts.get("cpu_line_png")

            # Memory
            mem_ts_timestamps, mem_avg = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.mem.usage", "avg",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            _, mem_max = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.mem.usage", "max",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            _, mem_min = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.mem.usage", "min",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            if mem_ts_timestamps and (mem_avg or mem_max or mem_min):
                min_len = min(len(mem_ts_timestamps), len(mem_avg), len(mem_max), len(mem_min))
                mem_ts_png = _timeseries_png(
                    mem_ts_timestamps[:min_len], mem_avg[:min_len], mem_max[:min_len], mem_min[:min_len],
                    title="Memory usage split by hosts", y_label="Memory usage %", legend_base="Memory usage %", tick_interval="1h"
                )
            else:
                mem_ts_png = charts.get("mem_line_png")

            # Disk
            disk_ts_timestamps, disk_avg = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.disk.usedPct", "avg",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            _, disk_max = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.disk.usedPct", "max",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            _, disk_min = _query_metric_timeseries(
                tenant_url, access_token, "builtin:host.disk.usedPct", "min",
                start_iso, end_iso, resolution="1m", entity_selector=entity_selector
            )
            if disk_ts_timestamps and (disk_avg or disk_max or disk_min):
                min_len = min(len(disk_ts_timestamps), len(disk_avg), len(disk_max), len(disk_min))
                disk_ts_png = _timeseries_png(
                    disk_ts_timestamps[:min_len], disk_avg[:min_len], disk_max[:min_len], disk_min[:min_len],
                    title="Disk usage split by hosts", y_label="Disk usage %", legend_base="Disk usage %", tick_interval="1h"
                )
            else:
                disk_ts_png = charts.get("disk_line_png")
        except Exception:
            cpu_ts_png = cpu_ts_png or charts.get("cpu_line_png")
            mem_ts_png = mem_ts_png or charts.get("mem_line_png")
            disk_ts_png = disk_ts_png or charts.get("disk_line_png")

        if action == "preview":
            def _to_datauri(png_bytes):
                return None if not png_bytes else "data:image/png;base64," + base64.b64encode(png_bytes).decode("ascii")
            charts_b64 = {k: _to_datauri(v) for k, v in charts.items() if v is not None}
            charts_b64["cpu_timeseries_png"] = _to_datauri(cpu_ts_png)
            charts_b64["mem_timeseries_png"] = _to_datauri(mem_ts_png)
            charts_b64["disk_timeseries_png"] = _to_datauri(disk_ts_png)
            toplists = charts.get("_toplists", {})
            return render(request, "daily_activity.html", {
                "validated": True, "tenant_url": tenant_url, "timeframe": timeframe,
                "management_zone": management_zone, "columns": columns, "table_rows": table_rows,
                "charts": charts_b64, "toplists": toplists,
            })

        # Resolve MZ display name for headings (optional)
        mz_display = management_zone_name or management_zone or "All"
        if mz_display and (re.fullmatch(r"-?\d+", mz_display) or re.fullmatch(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", mz_display)):
            try:
                zones = _fetch_management_zones(tenant_url, access_token)
                for z in zones:
                    if str(z.get("id")) == str(management_zone):
                        mz_display = z.get("name") or mz_display
                        break
            except Exception:
                pass

        # ===========================
        # Excel download (ONE sheet: table + charts)
        # ===========================
        if file_format == "excel" and action == "download":
            wb = Workbook()
            ws = wb.active
            ws.title = "Host Metrics"

            # ----- Title Section -----
            title = f"Capacity Management Report For {mz_display if mz_display else 'All'}"
            ws.merge_cells("A1:H4")
            ws["A1"].value = title
            ws["A1"].font = Font(size=18, bold=True)
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

            ws.merge_cells("A5:H5")
            ws["A5"].value = "Prepared by - KarvTech Team"
            ws["A5"].font = Font(size=12, italic=True, bold=True)
            ws["A5"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

            ws.merge_cells("A6:H6")
            ws["A6"].value = f"Date - {date.today().strftime('%d-%m-%Y')}"
            ws["A6"].font = Font(size=12, italic=True, bold=True)
            ws["A6"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

            # ----- Table Header -----
            start_row = 8
            headers = [
                "Host Name",
                "Host Availability %",
                "CPU Usage Max %",
                "CPU Usage Min %",
                "CPU Usage Avg %",
                "Memory Usage Max %",
                "Memory Usage Min %",
                "Memory Usage Avg %",
                "Disk Usage Max %",
                "Disk Usage Min %",
                "Disk Usage Avg %",
            ]
            thin = Side(border_style="thin", color="000000")

            for idx, col_name in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=idx, value=col_name)
                cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            def _fmt(v):
                if v is None or v == "": return ""
                try: return round(float(v), 2)
                except: return v

            def _fmt_percent(v):
                if v is None or v == "": return ""
                try: return f"{round(float(v), 2)}%"
                except: return v

            hosts_sorted = sorted(per_host.items(), key=lambda kv: (kv[1].get("Host Name") or "").lower())

            cpu_avg_key = "CPU Usage (%)"
            cpu_min_key = "CPU Usage (Min %)"
            cpu_max_key = "CPU Usage (Max %)"
            mem_avg_key = "Memory Usage (%)"
            mem_min_key = "Memory Usage (Min %)"
            mem_max_key = "Memory Usage (Max %)"
            disk_avg_key = "Disk Usage (%)"
            disk_min_key = "Disk Usage (Min %)"
            disk_max_key = "Disk Usage (Max %)"

            # ----- Write Table Rows -----
            for _, rec in hosts_sorted:
                row = [
                    rec.get("Host Name", ""),
                    _fmt(rec.get("Host Availability (%)", "")) / 100,  # percent format applied later
                    _fmt(rec.get(cpu_max_key, "")) / 100,
                    _fmt(rec.get(cpu_min_key, "")) / 100,
                    _fmt(rec.get(cpu_avg_key, "")) / 100,
                    _fmt(rec.get(mem_max_key, "")) / 100,
                    _fmt(rec.get(mem_min_key, "")) / 100,
                    _fmt(rec.get(mem_avg_key, "")) / 100,
                    _fmt(rec.get(disk_max_key, "")) / 100,
                    _fmt(rec.get(disk_min_key, "")) / 100,
                    _fmt(rec.get(disk_avg_key, "")) / 100,
                ]

                ws.append(row)
                r = ws.max_row

                # Apply borders
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # Apply percent formatting to ALL metric columns (B → K)
                for c in range(2, 12):  
                    ws.cell(row=r, column=c).number_format = "0.00%"   # show 2 decimal + % sign


            # Autofit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

            # ===========================
            # CHARTS ON SAME SHEET — STARTING ROW 14
            # ===========================
            def _embed(sheet, png_bytes, cell, width=950):
                try:
                    pil = PILImage.open(io.BytesIO(png_bytes))
                    if pil.mode not in ("RGB", "RGBA"):
                        pil = pil.convert("RGBA")
                    buff = io.BytesIO()
                    pil.save(buff, format="PNG")
                    buff.seek(0)
                    img = XLImage(buff)
                    img.width = width
                    img.height = int(width * (pil.height / pil.width))
                    sheet.add_image(img, cell)
                    rows_needed = int(img.height / 20) + 3
                    return rows_needed
                except:
                    return 30

            chart_row = 14  # <==== YOU REQUESTED THIS

            # CPU Chart
            if cpu_ts_png:
                used = _embed(ws, cpu_ts_png, f"A{chart_row}")
                chart_row += used

            # Memory Chart
            if mem_ts_png:
                used = _embed(ws, mem_ts_png, f"A{chart_row}")
                chart_row += used

            # Disk Chart
            if disk_ts_png:
                used = _embed(ws, disk_ts_png, f"A{chart_row}")
                chart_row += used

            # Save file
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="Capacity_Management.xlsx"'
            return resp


        # ===========================
        # PDF download
        # ===========================
        elif file_format == "pdf" and action == "download":
            buf = io.BytesIO()
            doc = SimpleDocTemplate(
                buf, pagesize=landscape(A4),
                leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30
            )
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('title_style', parent=styles['Title'],
                                         fontSize=22, leading=24, alignment=0, spaceAfter=6)
            meta_style = ParagraphStyle('meta_style', parent=styles['Normal'],
                                        fontSize=9, leading=11, alignment=0, italic=True)
            small_wrap = ParagraphStyle('small_wrap', parent=styles['Normal'],
                                        fontSize=7, leading=8, spaceBefore=0, spaceAfter=0)
            header_wrap = ParagraphStyle('header_wrap', parent=styles['Normal'],
                                         fontSize=8, leading=9, alignment=1)

            elements = []
            report_title = f"Capacity Management Report For {mz_display if mz_display else 'All'}"
            elements.append(
                Table(
                    [[Paragraph(f"<b>{report_title}</b>", title_style)],
                     [Paragraph("Prepared by - KarvTech Team", meta_style)],
                     [Paragraph(f"Date - {date.today().strftime('%d-%m-%Y')}", meta_style)]],
                    colWidths=[doc.width]
                )
            )
            elements.append(Spacer(1, 8))
            elements.append(Paragraph(f"Timeframe: {timeframe} (UTC: {start_iso} → {end_iso})", styles['Normal']))
            elements.append(Spacer(1, 8))

            # Remove Host ID
            host_id_index = None
            for i, c in enumerate(columns):
                if re.search(r'host\s*id', str(c), re.I):
                    host_id_index = i
                    break
            if host_id_index is not None:
                filtered_columns = [c for j, c in enumerate(columns) if j != host_id_index]
                filtered_rows = [[val for j, val in enumerate(row) if j != host_id_index] for row in table_rows]
            else:
                filtered_columns = columns[:]
                filtered_rows = [row[:] for row in table_rows]

            # Remove Host Availability Min/Max
            remove_patterns = re.compile(r'host\s*availability.*\b(min|max)\b', re.I)
            keep_index_pairs = [(i, c) for i, c in enumerate(filtered_columns) if not remove_patterns.search(str(c))]
            keep_indexes = [i for i, _ in keep_index_pairs]
            pdf_columns = [filtered_columns[i] for i in keep_indexes]
            pdf_table_rows = [[row[i] if i < len(row) else '' for i in keep_indexes] for row in filtered_rows]

            # Reorder: Host Name, Host Availability, then CPU/Mem/Disk usage, min, max
            indexed = list(enumerate(pdf_columns))
            ordered_indexes, ordered_columns = [], []

            # Host Name
            for i, h in indexed:
                if re.search(r'host\s*name', str(h), re.I):
                    ordered_indexes.append(i)
                    ordered_columns.append(h)
                    break

            # Host Availability (overall only)
            for i, h in indexed:
                if i in ordered_indexes:
                    continue
                if re.search(r'host\s*availability(?!.*\b(min|max)\b)', str(h), re.I):
                    ordered_indexes.append(i)
                    ordered_columns.append(h)
                    break

            metrics = [
                ('CPU', r'cpu'),
                ('Memory', r'\b(mem|memory)\b'),
                ('Disk', r'disk'),
            ]
            for _, metric_re in metrics:
                # Usage
                for i, h in indexed:
                    if i in ordered_indexes:
                        continue
                    if re.search(metric_re, str(h), re.I) and re.search(r'usage', str(h), re.I) and not re.search(r'\b(min|max)\b', str(h), re.I):
                        ordered_indexes.append(i)
                        ordered_columns.append(h)
                        break
                # Min
                for i, h in indexed:
                    if i in ordered_indexes:
                        continue
                    if re.search(metric_re, str(h), re.I) and re.search(r'\bmin\b', str(h), re.I):
                        ordered_indexes.append(i)
                        ordered_columns.append(h)
                        break
                # Max
                for i, h in indexed:
                    if i in ordered_indexes:
                        continue
                    if re.search(metric_re, str(h), re.I) and re.search(r'\bmax\b', str(h), re.I):
                        ordered_indexes.append(i)
                        ordered_columns.append(h)
                        break

            # Append remaining
            for i, h in indexed:
                if i not in ordered_indexes:
                    ordered_indexes.append(i)
                    ordered_columns.append(h)

            final_rows = [[row[i] if i < len(row) else '' for i in ordered_indexes] for row in pdf_table_rows]
            data = [ordered_columns] + final_rows

            # Convert to Paragraphs (after numeric-detection)
            for r in range(len(data)):
                for c in range(len(data[r])):
                    text = "" if data[r][c] is None else str(data[r][c])
                    if r == 0:
                        data[r][c] = Paragraph(f"<b>{text}</b>", header_wrap)
                    else:
                        data[r][c] = Paragraph(text, small_wrap)

            # Column widths
            num_cols = len(ordered_columns) if ordered_columns else 1
            rel_weights = [3.0] + [1.0] * (num_cols - 1)
            total_weight = sum(rel_weights)
            available_width = doc.width
            col_widths = [available_width * (w / total_weight) for w in rel_weights]

            # ensure first column wide enough
            min_first = 140
            if num_cols >= 1 and col_widths[0] < min_first:
                remaining = available_width - min_first
                if remaining < 0:
                    remaining = 30 * (num_cols - 1)
                    min_first = available_width - remaining
                col_widths[0] = min_first
                for i in range(1, num_cols):
                    col_widths[i] = max(30, remaining * (rel_weights[i] / (total_weight - rel_weights[0])))

            # numeric col detection (by header text)
            numeric_regex = re.compile(r'(cpu|mem|memory|disk|usage|min|max|%)', re.I)
            numeric_cols = []
            for i, h in enumerate(ordered_columns):
                if numeric_regex.search(str(h)):
                    numeric_cols.append(i)

            table = Table(data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
            tbl_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5597")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor("#666666")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ])

            if num_cols >= 1:
                tbl_style.add('ALIGN', (0, 1), (0, -1), 'LEFT')
            for idx in numeric_cols:
                tbl_style.add('ALIGN', (idx, 1), (idx, -1), 'RIGHT')
                tbl_style.add('ALIGN', (idx, 0), (idx, 0), 'CENTER')

            table.setStyle(tbl_style)
            elements.append(table)
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("<b>Visualizations</b>", styles['Heading2']))
            elements.append(Spacer(1, 6))

            # charts
            if cpu_ts_png:
                elements.append(Paragraph("CPU utilization split by hosts", styles['Heading3']))
                try:
                    pil_img = PILImage.open(io.BytesIO(cpu_ts_png))
                    orig_w, orig_h = pil_img.size
                    target_width = doc.width
                    ratio = (orig_h / float(orig_w)) if orig_w else 0.4
                    target_height = target_width * ratio
                    elements.append(RLImage(io.BytesIO(cpu_ts_png), width=target_width, height=target_height))
                except Exception:
                    elements.append(RLImage(io.BytesIO(cpu_ts_png), width=doc.width, height=3.4 * inch))
                elements.append(Spacer(1, 10))

            if mem_ts_png:
                elements.append(Paragraph("Memory utilization split by hosts", styles['Heading3']))
                try:
                    pil_img = PILImage.open(io.BytesIO(mem_ts_png))
                    orig_w, orig_h = pil_img.size
                    target_width = doc.width
                    ratio = (orig_h / float(orig_w)) if orig_w else 0.4
                    target_height = target_width * ratio
                    elements.append(RLImage(io.BytesIO(mem_ts_png), width=target_width, height=target_height))
                except Exception:
                    elements.append(RLImage(io.BytesIO(mem_ts_png), width=doc.width, height=3.4 * inch))
                elements.append(Spacer(1, 10))

            if disk_ts_png:
                elements.append(Paragraph("Disk utilization split by hosts", styles['Heading3']))
                try:
                    pil_img = PILImage.open(io.BytesIO(disk_ts_png))
                    orig_w, orig_h = pil_img.size
                    target_width = doc.width
                    ratio = (orig_h / float(orig_w)) if orig_w else 0.4
                    target_height = target_width * ratio
                    elements.append(RLImage(io.BytesIO(disk_ts_png), width=target_width, height=target_height))
                except Exception:
                    elements.append(RLImage(io.BytesIO(disk_ts_png), width=doc.width, height=3.4 * inch))
                elements.append(Spacer(1, 10))

            doc.build(elements)
            buf.seek(0)
            resp = HttpResponse(buf, content_type='application/pdf')
            resp["Content-Disposition"] = f'attachment; filename="Capacity_Management.pdf"'
            return resp

    # initial GET
    return render(request, "daily_activity.html")




########################################################################################################################################################
########################################################################################################################################################
########################################################################################################################################################




@login_required(login_url='login')
def user_management(request):
    from django.http import HttpResponse

    if request.method != "POST":
        return render(request, "user_management.html")

    # -------------------------------
    # Get Inputs
    # -------------------------------
    account_uuid = request.POST.get("account_uuid")
    client_id = request.POST.get("client_id")
    client_secret = request.POST.get("client_secret")
    report_format = request.POST.get("format")
    timeframe = request.POST.get("timeframe")

    if not account_uuid or not client_id or not client_secret:
        return render(request, "user_management.html", {"error": "All fields required."})

    try:
        # -----------------------------------------------------------------
        # Call the shared generator (NO duplication)
        # -----------------------------------------------------------------
        buf, mime, fname = generate_user_management_report_file(
            account_uuid=account_uuid,
            client_id=client_id,
            client_secret=client_secret,
            timeframe=timeframe,
            report_format=report_format
        )

        # -----------------------------------------------------------------
        # Return file to user
        # -----------------------------------------------------------------
        response = HttpResponse(buf.getvalue(), content_type=mime)
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        return response

    except Exception as e:
        return render(request, "user_management.html", {"error": str(e)})



def generate_user_management_report_file(account_uuid, client_id, client_secret, timeframe, report_format):
    """
    Builds the actual User Management report file and returns:
    (buffer, mime_type, filename)
    Used by both the scheduler and the manual download.
    """

    import io, csv, requests
    from datetime import datetime, timedelta, timezone
    from dateutil import parser
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch

    IAM_BASE = "https://api.dynatrace.com/iam/v1"

    # ---------------------------------------------------------------------
    # 1️⃣ Get OAuth Token
    # ---------------------------------------------------------------------
    token_url = "https://sso.dynatrace.com/sso/oauth2/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "resource": f"urn:dtaccount:{account_uuid}",
    }
    token_resp = requests.post(token_url, data=data)
    token_resp.raise_for_status()
    access_token = token_resp.json()["access_token"]

    # ---------------------------------------------------------------------
    # 2️⃣ Fetch users
    # ---------------------------------------------------------------------
    users_url = f"{IAM_BASE}/accounts/{account_uuid}/users"
    headers = {"Authorization": f"Bearer {access_token}"}
    users_resp = requests.get(users_url, headers=headers)
    users_resp.raise_for_status()
    users_data = users_resp.json().get("items", [])

    if not users_data:
        raise ValueError("No users found.")

    # -------------------------------
    # ✅ 3. TIMEFRAME FILTER
    # -------------------------------
    now = datetime.now(timezone.utc)
    today_start = datetime.combine(now.date(), datetime.min.time(), tzinfo=timezone.utc)
    today_end = datetime.combine(now.date(), datetime.max.time(), tzinfo=timezone.utc)

    if timeframe == "today":
        start_date, end_date = today_start, today_end
    elif timeframe == "yesterday":
        start_date = today_start - timedelta(days=1)
        end_date = today_end - timedelta(days=1)
    elif timeframe == "daily":
        start_date = now - timedelta(days=1)
        end_date = now
    elif timeframe == "weekly":
        start_date = now - timedelta(days=7)
        end_date = now
    elif timeframe == "monthly":
        start_date = now - timedelta(days=30)
        end_date = now
    else:
        start_date = now - timedelta(days=365)
        end_date = now

    filtered_users = []
    login_counts = {}

    for u in users_data:
        last_login = u.get("userLoginMetadata", {}).get("lastSuccessfulLogin")

        if last_login:
            try:
                dt = parser.isoparse(last_login)
                if start_date <= dt <= end_date:
                    filtered_users.append(u)
                    login_date = dt.date()
                    login_counts[login_date] = login_counts.get(login_date, 0) + 1
            except:
                pass

    if not filtered_users:
        filtered_users = users_data


    headers_row = [
        "UID", "Email", "First Name", "Last Name",
        "Status", "Emergency Contact", "Last Login", "Login Count"
    ]

    buf = io.BytesIO()
    filename = f"User_Management_{datetime.now().strftime('%Y%m%d_%H%M')}"
    mime = None

    # =====================================================================================
    # ✅ EXCEL EXPORT (SINGLE SHEET ONLY – Users + Summary Below)
    # =====================================================================================
    # =====================================================================================
    # ✅ EXCEL EXPORT (SINGLE SHEET — USERS + SUMMARY + CHART)
    # =====================================================================================
    if report_format.lower() == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        import io
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "User Management"

        # ---------------------------------------
        # HEADER BLOCK
        # ---------------------------------------
        ws.merge_cells("A1:H4")
        ws["A1"].value = "User Management Report"
        ws["A1"].font = Font(size=24, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A5:H5")
        ws["A5"].value = "Prepared by - KarvTech Team"
        ws["A5"].font = Font(size=12, bold=True)

        ws.merge_cells("A6:H6")
        ws["A6"].value = f"Date - {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
        ws["A6"].font = Font(size=12, bold=True)

        ws.append([])              # empty row
        ws.append(headers_row)      # USERS TABLE HEADERS (ROW 8)

        # ---------------------------------------
        # STYLE HEADER ROW
        # ---------------------------------------
        header_idx = ws.max_row
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[header_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # ---------------------------------------
        # USERS DATA ROWS
        # ---------------------------------------
        for u in filtered_users:
            last_login = u.get("userLoginMetadata", {}).get("lastSuccessfulLogin", "")
            login_count = 0
            if last_login:
                try:
                    dt = parser.isoparse(last_login)
                    login_count = login_counts.get(dt.date(), 0)
                except:
                    pass

            ws.append([
                u.get("uid"),
                u.get("email"),
                u.get("name"),
                u.get("surname"),
                u.get("userStatus"),
                u.get("emergencyContact"),
                last_login,
                login_count
            ])

        # ============================================
        # SUMMARY SECTION BELOW USERS TABLE
        # ============================================
        ws.append([])                               # gap
        ws.append(["Login Summary"])                # title
        ws["A" + str(ws.max_row)].font = Font(bold=True)

        ws.append(["Login Date", "Login Count"])    # header
        summary_header_idx = ws.max_row

        for cell in ws[summary_header_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # ============================================
        # SUMMARY DATA ROWS
        # ============================================
        total_logins = 0
        if login_counts:
            for login_date, count in sorted(login_counts.items()):
                ws.append([login_date.strftime("%Y-%m-%d"), count])
                total_logins += count
        else:
            ws.append(["No login data", 0])

        ws.append(["Total Logins", total_logins])

        # ============================================
        # CHART BELOW SUMMARY
        # ============================================
        if login_counts:
            chart = BarChart()
            chart.title = "User Login Count by Date"

            # location of summary header
            summary_start = summary_header_idx

            # data range (Login Count)
            data = Reference(ws, min_col=2, min_row=summary_start,
                            max_row=ws.max_row - 1)

            # labels range (Login Date)
            labels = Reference(ws, min_col=1, min_row=summary_start + 1,
                            max_row=ws.max_row - 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)

            # place the chart
            ws.add_chart(chart, f"D{summary_start}")

        # ============================================
        # AUTO-FIT COLUMNS
        # ============================================
        for i, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        # ============================================
        # SAVE TO BUFFER
        # ============================================
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        buf = excel_buffer
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        fname = f"User_Management_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


    # ---------------------------------------------------------------------
    # 📗 CSV
    # ---------------------------------------------------------------------
    elif report_format.lower() == "csv":
        text_buf = io.StringIO()
        writer = csv.writer(text_buf)
        writer.writerow(headers_row)
        for u in filtered_users:
            writer.writerow([
                u.get("uid"), u.get("email"), u.get("name"), u.get("surname"),
                u.get("userStatus"), u.get("emergencyContact"),
                u.get("userLoginMetadata", {}).get("lastSuccessfulLogin", ""), ""
            ])
        buf = io.BytesIO(text_buf.getvalue().encode("utf-8"))
        mime = "text/csv"
        fname = f"{filename}.csv"

    # ---------------------------------------------------------------------
    # 📕 ADVANCED PDF (HEADER + FULL TABLE + SUMMARY + CHART)
    # ---------------------------------------------------------------------
    elif report_format.lower() == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, PageBreak
        )
        from reportlab.lib.units import inch
        from reportlab.graphics.shapes import Drawing, Rect, String
        import io

        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(A4),
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name="TitleStyle",
            fontSize=20,
            leading=24,
            alignment=1,
            spaceAfter=12,
            fontName="Helvetica-Bold"
        ))

        styles.add(ParagraphStyle(
            name="NormalStyle",
            fontSize=12,
            leading=16,
            spaceAfter=6,
            fontName="Helvetica"
        ))

        elements = []

        # ---------------------------------------------------------------
        # HEADER (SAME AS EXCEL)
        # ---------------------------------------------------------------
        elements.append(Paragraph("User Management Report", styles["TitleStyle"]))
        elements.append(Paragraph("Prepared by - KarvTech Team", styles["NormalStyle"]))
        elements.append(Paragraph(
            f"Date - {datetime.now().strftime('%d-%b-%Y %I:%M %p')}",
            styles["NormalStyle"]
        ))
        elements.append(Spacer(1, 0.4 * inch))

        # ---------------------------------------------------------------
        # USERS TABLE
        # ---------------------------------------------------------------
        table_data = [headers_row]

        for u in filtered_users:
            last_login = u.get("userLoginMetadata", {}).get("lastSuccessfulLogin", "")
            login_count = 0
            if last_login:
                try:
                    dt = parser.isoparse(last_login)
                    login_count = login_counts.get(dt.date(), 0)
                except:
                    pass

            table_data.append([
                u.get("uid"),
                u.get("email"),
                u.get("name"),
                u.get("surname"),
                u.get("userStatus"),
                u.get("emergencyContact"),
                last_login,
                str(login_count)
            ])

        pdf_table = Table(table_data, repeatRows=1)
        pdf_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]))

        elements.append(pdf_table)

        # ---------------------------------------------------------------
        # PAGE BREAK
        # ---------------------------------------------------------------
        elements.append(PageBreak())

        # ---------------------------------------------------------------
        # SUMMARY TABLE
        # ---------------------------------------------------------------
        elements.append(Paragraph("Login Summary", styles["TitleStyle"]))

        summary_data = [["Login Date", "Login Count"]]

        total_logins = 0
        if login_counts:
            for login_date, count in sorted(login_counts.items()):
                summary_data.append([login_date.strftime("%Y-%m-%d"), str(count)])
                total_logins += count
            summary_data.append(["Total Logins", str(total_logins)])
        else:
            summary_data.append(["No Login Data", "0"])

        summary_table = Table(summary_data, repeatRows=1)
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # ---------------------------------------------------------------
        # BAR CHART (FULL CUSTOM CHART)
        # ---------------------------------------------------------------
        if login_counts:
            sorted_items = sorted(login_counts.items())
            labels = [d.strftime("%Y-%m-%d") for d, _ in sorted_items]
            values = [int(v) for _, v in sorted_items]

            drawing_width = 800
            drawing_height = 300
            drawing = Drawing(drawing_width, drawing_height)

            left_margin = 60
            bottom_margin = 60
            plot_width = drawing_width - 120
            plot_height = drawing_height - 120

            max_val = max(values) or 1

            gap = 6
            bar_width = max(20, (plot_width - gap * len(values)) / len(values))

            palette = [
                colors.HexColor("#4CAF50"),
                colors.HexColor("#2196F3"),
                colors.HexColor("#FF9800"),
                colors.HexColor("#E91E63"),
                colors.HexColor("#9C27B0"),
            ]

            drawing.add(String(drawing_width / 2, drawing_height - 20,
                            "User Login Count by Date",
                            fontSize=14, textAnchor="middle"))

            x = left_margin
            for i, val in enumerate(values):
                height = (val / max_val) * plot_height
                drawing.add(Rect(x, bottom_margin, bar_width, height,
                                fillColor=palette[i % len(palette)],
                                strokeColor=colors.black))

                drawing.add(String(x + bar_width / 2, bottom_margin - 12,
                                labels[i], fontSize=6, textAnchor="middle"))

                drawing.add(String(x + bar_width / 2, bottom_margin + height + 5,
                                str(val), fontSize=7, textAnchor="middle"))

                x += bar_width + gap

            elements.append(drawing)

        # ---------------------------------------------------------------
        # FINISH PDF
        # ---------------------------------------------------------------
        doc.build(elements)
        pdf_buffer.seek(0)

        buf = pdf_buffer
        mime = "application/pdf"
        fname = f"{filename}.pdf"



    buf.seek(0)
    return buf, mime, fname



def generate_user_management_report(account_uuid, client_id, client_secret, timeframe, report_format, email):
    from django.core.mail import EmailMessage
    from django.conf import settings

    try:
        buf, mime, fname = generate_user_management_report_file(account_uuid, client_id, client_secret, timeframe, report_format)

        subject = f"User Management Report ({report_format.upper()})"
        body = (
            "Hello Team,\n\n"
            "Please find attached User Management Report.\n"
            f"Format: {report_format.upper()}\n\n"
            "Regards,\nKarvTech Automated Reporting System"
        )

        mail = EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, [email])
        mail.attach(fname, buf.getvalue(), mime)
        mail.send()
        logger.info(f"✅ User Management emailed to {email}")

    except Exception as e:
        logger.info(f"❌ Failed to send User Management report: {e}")



########################################################################################################################################################
########################################################################################################################################################
########################################################################################################################################################
# =========================
# SBOM HELPERS + VIEW + SCHEDULER ENTRYPOINT
# =========================
import json as _json

def _sbom_convert_to_ist(timestamp_ms):
    if not timestamp_ms:
        return ""
    try:
        utc_time = datetime.fromtimestamp(timestamp_ms / 1000, tz=dt_tz.utc)
        ist = dt_tz(timedelta(hours=5, minutes=30))
        return utc_time.astimezone(ist).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(timestamp_ms)

def _sbom_fetch_management_zones(tenant_url, token):
    # reuse your existing _fetch_management_zones if present
    return _fetch_management_zones(tenant_url, token)

def _sbom_fetch_entities_by_zone(tenant_url, token, zone_name):
    headers = {"Authorization": f"Api-Token {token}"}
    url = f'{tenant_url}/api/v2/entities?pageSize=500&entitySelector=mzName("{zone_name}"),type("PROCESS_GROUP_INSTANCE")'
    r = requests.get(url, headers=headers, timeout=60)
    r.raise_for_status()
    return r.json().get("entities", [])

def _sbom_fetch_entity_details(tenant_url, token, entity_id):
    headers = {"Authorization": f"Api-Token {token}"}
    url = f"{tenant_url}/api/v2/entities/{entity_id}"
    r = requests.get(url, headers=headers, timeout=60)
    if r.status_code != 200:
        return None
    return r.json()

def _sbom_build_dataset(tenant_url, token, zone_name):
    """
    Returns (entities_rows, tech_type_counts) where entities_rows is a list of dicts
    prepared for rendering.
    """
    base = _sbom_fetch_entities_by_zone(tenant_url, token, zone_name)
    rows = []
    tech_type_count = {}

    for e in base:
        entity_id = e.get("entityId")
        if not entity_id:
            continue

        detail = _sbom_fetch_entity_details(tenant_url, token, entity_id)
        if not detail:
            continue

        props = detail.get("properties", {})
        software_techs = props.get("softwareTechnologies", [])

        for tech in software_techs:
            ttype = tech.get("type")
            if ttype:
                tech_type_count[ttype] = tech_type_count.get(ttype, 0) + 1

        rows.append({
            "entityId": detail.get("entityId", ""),
            "type": detail.get("type", ""),
            "displayName": detail.get("displayName", ""),
            "firstSeenTms": _sbom_convert_to_ist(detail.get("firstSeenTms")),
            "lastSeenTms": _sbom_convert_to_ist(detail.get("lastSeenTms")),
            "detectedName": props.get("detectedName", ""),
            "bitness": props.get("bitness", ""),
            "metadata": props.get("metadata", []),                    # keep native, render later
            "softwareTechnologies": software_techs,                   # keep native, render later
            "listenPorts": props.get("listenPorts", []),
            "tags": detail.get("tags", []),
            "managementZones": detail.get("managementZones", []),
        })

    return rows, tech_type_count

def _render_sbom_report_bytes(entities, tech_counts, fmt, selected_zone):
    """
    Render SBOM as BytesIO for excel / csv / json.
    Return (buf, mime, filename)
    """
    fmt = (fmt or "").lower()

    # ---------------- CSV ----------------
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        cols = [
            "entityId", "displayName", "firstSeenTms", "lastSeenTms",
            "detectedName", "metadata", "softwareTechnologies"
        ]
        writer.writerow(cols)
        for row in entities:
            writer.writerow([
                row.get("entityId", ""),
                row.get("displayName", ""),
                row.get("firstSeenTms", ""),
                row.get("lastSeenTms", ""),
                row.get("detectedName", ""),
                _json.dumps(row.get("metadata", []), ensure_ascii=False),
                _json.dumps(row.get("softwareTechnologies", []), ensure_ascii=False),
            ])
        out = io.BytesIO(buf.getvalue().encode("utf-8"))
        out.seek(0)
        return out, "text/csv", f"Software_Inventory_Report.csv"

    # ---------------- JSON ----------------
    if fmt == "json":
        org = "COSMOS BANK"
        packages_list = []
        for row in entities:
            techs = row.get("softwareTechnologies", [])
            for t in techs:
                pkg = {
                    "name": t.get("type", ""),
                    "SPDXID": f"SPDX-Package-{t.get('type','')}",
                    "versionInfo": t.get("version", ""),
                    "licenseDeclared": t.get("edition", ""),
                    "downloadLocation": "C:/Users/Downloads/",
                    "filesAnalyzed": False
                }
                packages_list.append(pkg)

        result = {
            "SPDXID": 1,
            "name": "KarvOps",
            "creationInfo": {
                "created": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
                "creators": [f"Organization: {org}", "Tool: KarvOps-v1.0"]
            },
            "packages": packages_list
        }
        js = _json.dumps(result, indent=4)
        out = io.BytesIO(js.encode("utf-8"))
        out.seek(0)
        return out, "application/json", f"SOftware_Inventory_Report.json"

    # ---------------- EXCEL ----------------
    if fmt == "excel":
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"SBHOM_{selected_zone}"

        # Title section
        from datetime import date as _date
        title = f"Software Inventory Report for {selected_zone}"
        ws.merge_cells("A1:H4")
        ws["A1"].value = title
        ws["A1"].font = Font(size=24, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

        ws.merge_cells("A5:H5")
        ws["A5"].value = "Prepared by - KarvTech Team"
        ws["A5"].font = Font(size=12, italic=True, bold=True)
        ws["A5"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

        ws.merge_cells("A6:H6")
        ws["A6"].value = f"Date - {_date.today().strftime('%d-%m-%Y')}"
        ws["A6"].font = Font(size=12, italic=True, bold=True)
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

        start_row = 8

        # Headers
        headers = [
            "Entity ID", "Display Name", "First Seen (IST)",
            "Last Seen (IST)", "Detected Name", "Metadata",
            "Software Technologies"
        ]
        ws.append(["" for _ in headers])  # keep your spacing
        ws.append(headers)
        header_row = start_row
        for idx, col_name in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=idx, value=col_name)
            c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row in entities:
            ws.append([
                row.get("entityId", ""),
                row.get("displayName", ""),
                row.get("firstSeenTms", ""),
                row.get("lastSeenTms", ""),
                row.get("detectedName", ""),
                _json.dumps(row.get("metadata", []), ensure_ascii=False),
                _json.dumps(row.get("softwareTechnologies", []), ensure_ascii=False),
            ])

        # Autofit
        for col in ws.columns:
            max_len = 0
            idx = col[0].column
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)

        # Summary table (type → versions, count)
        header_font = Font(color="000000", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

        summary_start = ws.max_row + 3
        ws.cell(row=summary_start, column=1, value="Software Technology Type")
        ws.cell(row=summary_start, column=2, value="Version(s)")
        ws.cell(row=summary_start, column=3, value="Count")
        for col in range(1, 4):
            c = ws.cell(row=summary_start, column=col)
            c.font = header_font
            c.alignment = center_align
            c.border = thin_border

        # Build type→versions/count
        type_info = {}
        for row in entities:
            techs = row.get("softwareTechnologies", [])
            for t in techs:
                t_type = str(t.get("type", "")).strip()
                t_version = str(t.get("version", "")).strip() or "-"
                if not t_type:
                    continue
                if t_type not in type_info:
                    type_info[t_type] = {"versions": set(), "count": 0}
                type_info[t_type]["versions"].add(t_version)
                type_info[t_type]["count"] += 1

        cur = summary_start + 1
        for t_type, info in type_info.items():
            versions_str = ",  ".join(sorted(v for v in info["versions"] if v != "-")) or "-"
            ws.cell(row=cur, column=1, value=t_type)
            ws.cell(row=cur, column=2, value=versions_str)
            ws.cell(row=cur, column=3, value=int(info["count"]))
            for col in range(1, 4):
                c = ws.cell(row=cur, column=col)
                c.alignment = center_align
                c.border = thin_border
                if cur % 2 == 0:
                    c.fill = alt_fill
            cur += 1

        # Bar chart
        if type_info:
            chart = BarChart()
            chart.title = "Software Technologies by Type"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Technology Type"
            chart.width = 25
            chart.height = 15

            data_ref = Reference(ws, min_col=3, min_row=summary_start + 1, max_row=cur - 1)
            cats_ref = Reference(ws, min_col=1, min_row=summary_start + 1, max_row=cur - 1)
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cats_ref)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showSerName = False
            chart.dataLabels.showCatName = False
 

            ws.add_chart(chart, f"E{summary_start}")

        # Sheet 2 (SBHOM_Details)
        ws2 = wb.create_sheet(title="SBHOM_Details")
        sheet2_headers = [
            "SR.No", "Name of Software", "Created Date", "Name Of Organization",
            "Tool", "Name of Package", "SPDXID", "Version",
            "License Declared", "Download Location", "FilesAnalyzed"
        ]
        ws2.append(sheet2_headers)
        for col in range(1, len(sheet2_headers) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        organization = "COSMOS Bank"
        sr_no = 1
        today = datetime.now().strftime("%d-%m-%Y")

        for row in entities:
            techs = row.get("softwareTechnologies", [])
            for t in techs:
                pkg_type = t.get("type", "-")
                version_info = t.get("version", "-")
                license_declared = t.get("edition", "-")
                ws2.append([
                    sr_no, "KarvOps", today, organization,
                    "KarvOps-v1.0", pkg_type, f"SPDX-Package-{pkg_type}",
                    version_info, license_declared, "C:/Users/Downloads/", False
                ])
                sr_no += 1

        for col in ws2.columns:
            max_len = 0
            idx = col[0].column
            for c in col:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
            ws2.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"Software_Inventory_Report.xlsx"

    raise ValueError(f"Unsupported SBOM format: {fmt}")

@login_required(login_url='login')
def sbom(request):
    tenant_url = (request.POST.get("tenant_url") or request.GET.get("tenant_url") or "").strip().rstrip("/")
    if "apps.dynatrace.com" in tenant_url:
        tenant_url = tenant_url.replace("apps.dynatrace.com", "live.dynatrace.com")
    token = (request.POST.get("access_token") or request.GET.get("access_token") or "").strip()
    action = request.POST.get("action")
    file_format = request.GET.get("format") or request.POST.get("format")
    selected_zone = request.POST.get("management_zone")

    if not tenant_url or not token:
        return render(request, "sbom.html", {"error": "Please provide tenant URL and token."})

    # Step 1: validate → fetch zones
    if request.method == "POST" and action == "validate":
        try:
            zones = _sbom_fetch_management_zones(tenant_url, token)
            return render(request, "sbom.html", {
                "validated": True,
                "tenant_url": tenant_url,
                "access_token": token,
                "zones": zones
            })
        except Exception as e:
            return render(request, "sbom.html", {"error": str(e)})

    # Step 2: download
    if request.method == "POST" and action == "download":
        if not selected_zone:
            return render(request, "sbom.html", {"error": "Please select a Management Zone."})
        try:
            entities, tech_counts = _sbom_build_dataset(tenant_url, token, selected_zone)
            if not entities:
                return render(request, "sbom.html", {"error": f"No entities found for {selected_zone}."})
            buf, mime, fname = _render_sbom_report_bytes(entities, tech_counts, file_format, selected_zone)
            resp = HttpResponse(buf.getvalue(), content_type=mime)
            resp["Content-Disposition"] = f'attachment; filename="{fname}"'
            return resp
        except Exception as e:
            return render(request, "sbom.html", {"error": str(e)})

    return render(request, "sbom.html")

def generate_sbom_report(tenant_url, access_token, management_zone, report_format, email):
    """
    Called by scheduler. Builds SBOM dataset, renders requested format, and emails it.
    """
    try:
        tenant_url = _normalize_tenant(tenant_url)
        entities, tech_counts = _sbom_build_dataset(tenant_url, access_token, management_zone)
        if not entities:
            logger.info("⚠️ SBOM: no entities found. Skipping email.")
            return
        buf, mime, fname = _render_sbom_report_bytes(entities, tech_counts, report_format, management_zone)

        subject = f"SBOM Report - {management_zone}"
        body = (
            "Hello Team,\n\n"
            "Please find attached Software Inventory Report.\n"
            f"Management Zone: {management_zone}\n"
            f"Format: {report_format.upper()}\n\n"
            "Regards,\nKarvTech Automated Reporting System"
        )
        email_msg = EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, [email])
        email_msg.attach(fname, buf.getvalue(), mime)
        email_msg.send(fail_silently=False)
        logger.info(f"✅ SBOM emailed to {email}")
    except Exception as e:
        logger.info(f"❌ SBOM scheduler error: {e}")


##################################################################################################################################################################################################################
##################################################################################################################################################################################################################
##################################################################################################################################################################################################################
 ########Capacity###########
CPU_THRESHOLD = 75.0
MEM_THRESHOLD = 75.0
 
# =====================================================
# Dynatrace API helper
# =====================================================
def dt_get(tenant_url, token, path, params=None):
    r = requests.get(
        f"{tenant_url}{path}",
        headers={"Authorization": f"Api-Token {token}"},
        params=params,
        timeout=30,
    )
    r.raise_for_status()
    return r.json()
 
# =====================================================
# Timeframe + Resolution (VALID FOR DT API)
# =====================================================
def resolve_timeframe(tf):
    now = datetime.utcnow()
 
    if tf == "today":
        return now.replace(hour=0, minute=0, second=0), now, "1m"
 
    if tf == "yesterday":
        start = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0)
        return start, start + timedelta(days=1), "1m"
 
    if tf == "24h":
        return now - timedelta(hours=24), now, "1m"
 
    if tf == "7d":
        return now - timedelta(days=7), now, "1m"
 
    if tf == "30d":
        return now - timedelta(days=30), now, "1h"
 
    return now - timedelta(days=7), now, "1m"
 
# =====================================================
# Host cache
# =====================================================
HOST_CACHE = {}
 
def get_host_details(tenant, token, host_id):
    if host_id in HOST_CACHE:
        return HOST_CACHE[host_id]
 
    data = dt_get(tenant, token, f"/api/v2/entities/{host_id}")
    props = data.get("properties", {})
 
    name = data.get("displayName") or props.get("hostname") or host_id
 
    ip = (
        props.get("primaryIpAddress")
        or props.get("ipAddress")
        or props.get("detectedIpAddresses")
    )
 
    if isinstance(ip, list):
        ip = ", ".join(str(x) for x in ip)
    elif ip is None:
        ip = "NA"
 
    cores = props.get("cpuCores", "NA")
 
    HOST_CACHE[host_id] = (name, ip, cores)
    return HOST_CACHE[host_id]
 
def get_disk_mount(tenant, token, disk_id):
    try:
        data = dt_get(tenant, token, f"/api/v2/entities/{disk_id}")
        return (
            data.get("properties", {}).get("mountPoint")
            or data.get("displayName")
            or disk_id
        )
    except:
        return disk_id
 
# =====================================================
# Helpers
# =====================================================
def latest_value(values):
    for v in reversed(values or []):
        if v is not None:
            return v
    return None
 
def pct(v):
    return f"{round(v, 2)} %"
 
def format_bytes(v):
    if v is None:
        return ""
    v = float(v)
    if v >= 1024**4:
        return f"{v / 1024**4:.2f} TiB"
    if v >= 1024**3:
        return f"{v / 1024**3:.2f} GiB"
    return f"{v / 1024**2:.2f} MiB"
 
 
 
# =====================================================
# Excel Writer
# =====================================================
def write_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    header = PatternFill("solid", fgColor="305496")
    border = Border(*(Side(style="thin"),) * 4)
    center = Alignment(horizontal="center")
 
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header
        cell.border = border
        cell.alignment = center
 
    for r, row in df.iterrows():
        for c, col in enumerate(df.columns, 1):
            ws.cell(row=r + 2, column=c, value=row[col]).border = border
 
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 28
 
# =====================================================
# MAIN VIEW
# =====================================================
import pandas as pd
from io import BytesIO
def capacity_management(request):
    context = {}
 
    if request.method == "POST":
        action = request.POST.get("action")
        tenant = request.POST.get("tenant_url")
        token = request.POST.get("access_token")
 
        # ---------------- VALIDATE ----------------
        if action == "validate":
            try:
                zones = dt_get(tenant, token, "/api/config/v1/managementZones").get("values", [])
                context.update({
                    "validated": True,
                    "tenant_url": tenant,
                    "access_token": token,
                    "zones": zones,
                    "message": "Tenant and token validated successfully",
                })
            except Exception as e:
                context["error"] = f"Validation failed: {e}"
 
        # ---------------- DOWNLOAD ----------------
        elif action == "download":
            timeframe = request.POST.get("timeframe")
            start, end, resolution = resolve_timeframe(timeframe)
 
            params = {
                "from": start.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "to": end.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "resolution": resolution,
            }
 
            # ================= CPU =================
            cpu_raw = dt_get(tenant, token, "/api/v2/metrics/query",
                             {"metricSelector": "builtin:host.cpu.usage", **params})
 
            cpu_map = {}
            for r in cpu_raw.get("result", []):
                for s in r.get("data", []):
                    h = s["dimensionMap"]["dt.entity.host"]
                    cpu_map.setdefault(h, []).extend(v for v in s["values"] if v is not None)
 
            cpu_rows = []
            for h, vals in cpu_map.items():
                name, ip, cores = get_host_details(tenant, token, h)
                cpu_rows.append({
                    "Host Name": name,
                    "Host IP": ip,
                    "CPU Cores": cores,
                    "CPU Usage Avg %": pct(sum(vals) / len(vals)),
                    "CPU > 75% Count": sum(1 for v in vals if v > CPU_THRESHOLD),
                })
 
            cpu_df = pd.DataFrame(cpu_rows)
 
            # ================= MEMORY =================
            mem_usage_raw = dt_get(
                tenant, token,
                "/api/v2/metrics/query",
                {"metricSelector": "builtin:host.mem.usage", **params}
            )
 
            mem_total_raw = dt_get(
                tenant, token,
                "/api/v2/metrics/query",
                {"metricSelector": "builtin:host.mem.total", **params}
            )
 
            mem_usage_map = {}
            mem_total_map = {}
 
            # --- Memory usage %
            for r in mem_usage_raw.get("result", []):
                for s in r.get("data", []):
                    h = s["dimensionMap"]["dt.entity.host"]
                    mem_usage_map.setdefault(h, []).extend(
                        v for v in s["values"] if v is not None
                    )
 
            # --- Total installed memory
            for r in mem_total_raw.get("result", []):
                for s in r.get("data", []):
                    h = s["dimensionMap"]["dt.entity.host"]
                    latest = latest_value(s.get("values"))
                    if latest is not None:
                        mem_total_map[h] = latest
 
            mem_rows = []
 
            for h, usage_vals in mem_usage_map.items():
                name, ip, _ = get_host_details(tenant, token, h)
 
                mem_rows.append({
                    "Host Name": name,
                    "Host IP": ip,
                    "Average Memory Usage %": pct(sum(usage_vals) / len(usage_vals)),
                    "Memory > 75% Count": sum(1 for v in usage_vals if v > MEM_THRESHOLD),
                    "Total Installed Memory": format_bytes(mem_total_map.get(h, 0)),
                })
 
            mem_df = pd.DataFrame(mem_rows).sort_values("Host Name")
 
 
            # ================= DISK (FULL) =================
            disk_metrics = {
                "Disk Used %": 'builtin:host.disk.usedPct:splitBy("dt.entity.host","dt.entity.disk")',
                "Disk Availability %": 'builtin:host.disk.free:splitBy("dt.entity.host","dt.entity.disk")',
                "Disk Used": 'builtin:host.disk.used:splitBy("dt.entity.host","dt.entity.disk")',
                "Disk Available": 'builtin:host.disk.avail:splitBy("dt.entity.host","dt.entity.disk")',
            }
 
            disk_data = {}
 
            for col, selector in disk_metrics.items():
                payload = dt_get(tenant, token, "/api/v2/metrics/query",
                                 {"metricSelector": selector, **params})
 
                for r in payload.get("result", []):
                    for s in r.get("data", []):
                        dm = s.get("dimensionMap", {})
                        host_id = dm.get("dt.entity.host")
                        disk_id = dm.get("dt.entity.disk")
 
                        val = latest_value(s.get("values"))
                        if val is None:
                            continue
 
                        key = (host_id, disk_id)
                        if key not in disk_data:
                            host_name, _, _ = get_host_details(tenant, token, host_id)
                            disk_data[key] = {
                                "Host": host_name,
                                "Mount Point": get_disk_mount(tenant, token, disk_id),
                            }
 
                        disk_data[key][col] = val
 
            disk_rows = []
            for r in disk_data.values():
                used = r.get("Disk Used", 0)
                avail = r.get("Disk Available", 0)
 
                r["Disk Used %"] = pct(r.get("Disk Used %", 0))
                r["Disk Availability %"] = pct(r.get("Disk Availability %", 0))
                r["Disk Used"] = format_bytes(used)
                r["Disk Available"] = format_bytes(avail)
                r["Total Disk"] = format_bytes(used + avail)
 
                disk_rows.append(r)
 
            disk_df = pd.DataFrame(disk_rows).sort_values(["Host", "Mount Point"])
 
            # ================= EXCEL =================
            wb = Workbook()
            wb.remove(wb.active)
 
            write_sheet(wb, "Disk Utilization", disk_df)
            write_sheet(wb, "Memory Utilization", mem_df)
            write_sheet(wb, "CPU Utilization", cpu_df)
 
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
 
            response = HttpResponse(
                buffer,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=Capacity_Report.xlsx"
            return response
 
    return render(request, "capacity_report.html", context)
 

def capacity_base(request):
    return render(request, "capacity_base.html") 

######################Predictive AI #################################
def ask_ai(request):
    return render(request, "ask_ai.html")
def predictive_ui(request):
    return render(request, "predictive_ui.html")
def generative_ui(request):
    return render(request, "generative_ui.html")

from datetime import date
from django.http import JsonResponse
from .models import Host

# use YOUR existing helpers:
# _query_metric_split_by_host
# get_real_host_name

TENANT_URL = "https://kdqxxxx.live.dynatrace.com"
API_TOKEN = ""

def HostLevelMetrics(request):
    total_saved = 0

    for i in range(0, 11):  # last 7 days
        day = date.today() - timedelta(days=i)

        start_time = f"now-{i}d/d"
        end_time = f"now-{i-1}d/d" if i > 1 else "now/d"

        availability = _query_metric_split_by_host(
            TENANT_URL, API_TOKEN,
            "builtin:host.availability", "avg",
            start_time, end_time
        )

        cpu = _query_metric_split_by_host(
            TENANT_URL, API_TOKEN,
            "builtin:host.cpu.usage", "avg",
            start_time, end_time
        )

        memory = _query_metric_split_by_host(
            TENANT_URL, API_TOKEN,
            "builtin:host.mem.usage", "avg",
            start_time, end_time
        )

        avail_map = {h["host_id"]: h for h in availability}
        cpu_map = {h["host_id"]: h for h in cpu}
        mem_map = {h["host_id"]: h for h in memory}

        all_hosts = set(avail_map) | set(cpu_map) | set(mem_map)

        for host_id in all_hosts:
            host_name = (
                avail_map.get(host_id, {}).get("host_name")
                or cpu_map.get(host_id, {}).get("host_name")
                or mem_map.get(host_id, {}).get("host_name")
                or host_id
            )

            Host.objects.update_or_create(
                host_name=host_name,
                record_date=day,
                defaults={
                    "host_availability": round(avail_map.get(host_id, {}).get("value_avg", 0), 2),
                    "cpu_usage": round(cpu_map.get(host_id, {}).get("value_avg", 0), 2),
                    "memory_usage": round(mem_map.get(host_id, {}).get("value_avg", 0), 2),
                }
            )

            total_saved += 1

    return JsonResponse({"status": "success", "hosts_saved": total_saved})

    ################Host graph#################

def _query_metric_split_by_host(tenant_url, api_token, metric_key, aggregation, start_time, end_time):
    url = f"{tenant_url.rstrip('/')}/api/v2/metrics/query"

    metric_selector = f'{metric_key}:{aggregation}:splitBy("dt.entity.host")'

    params = {
        "metricSelector": metric_selector,
        "from": start_time,
        "to": end_time,
        "resolution": "1h",
        "pageSize": 4000
    }

    headers = {
        "Authorization": f"Api-Token {api_token}",
        "Accept": "application/json"
    }

    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()

    data = response.json()
    results = []

    for res in data.get("result", []):
        for series in res.get("data", []):
            host_id = series["dimensions"][0]
            values = [v for v in series.get("values", []) if v is not None]

            if not values:
                continue

            avg_value = sum(values) / len(values)

            results.append({
                "host_id": host_id,
                "host_name": get_real_host_name(tenant_url, api_token, host_id),
                "value_avg": avg_value
            })

    return results

####################Service Table########################
from datetime import date
from django.http import JsonResponse
from .models import Service
TENANT_URL2 = "https://kdqxxx.live.dynatrace.com"
API_TOKEN2 = ""


from datetime import date, timedelta
from django.http import JsonResponse
from .models import Service

def ServiceLevelMetrics(request):
    total_saved = 0

    for i in range(5, 10):  # last 7 days
        day = date.today() - timedelta(days=i)

        start_time = f"now-{i}d/d"
        end_time = f"now-{i-1}d/d" if i > 1 else "now/d"

        requests_data = query_service_metric(TENANT_URL2, API_TOKEN2,
                                             "builtin:service.requestCount.total",
                                             None, start_time, end_time)

        response_time_data = query_service_metric(TENANT_URL2, API_TOKEN2,
                                                  "builtin:service.response.time",
                                                  "avg", start_time, end_time)

        error_rate_data = query_service_metric(TENANT_URL2, API_TOKEN2,
                                               "builtin:service.errors.total.rate",
                                               "avg", start_time, end_time)

        req_map = {s["service_id"]: s for s in requests_data}
        res_map = {s["service_id"]: s for s in response_time_data}
        err_map = {s["service_id"]: s for s in error_rate_data}

        all_services = set(req_map) | set(res_map) | set(err_map)

        for service_id in all_services:
            service_name = (
                req_map.get(service_id, {}).get("service_name")
                or res_map.get(service_id, {}).get("service_name")
                or err_map.get(service_id, {}).get("service_name")
                or service_id
            )

            Service.objects.update_or_create(
                service_name=service_name,
                record_date=day,
                defaults={
                    "request_count": round(req_map.get(service_id, {}).get("value", 0), 2),
                    "response_time": round(res_map.get(service_id, {}).get("value", 0), 2),
                    "failure_rate": round(err_map.get(service_id, {}).get("value", 0), 2),
                }
            )
            total_saved += 1

    return JsonResponse({"status": "success", "services_saved": total_saved})


import requests

def query_service_metric(tenant_url, api_token, metric_key, aggregation=None, start_time="now-1d", end_time="now"):
    url = f"{tenant_url.rstrip('/')}/api/v2/metrics/query"

    if aggregation:
        metric_selector = f'{metric_key}:{aggregation}:splitBy("dt.entity.service")'
    else:
        metric_selector = f'{metric_key}:splitBy("dt.entity.service")'

    params = {
        "metricSelector": metric_selector,
        "from": start_time,
        "to": end_time,
        "resolution": "1h",
        "pageSize": 4000
    }

    headers = {
        "Authorization": f"Api-Token {api_token}",
        "Accept": "application/json"
    }

    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()

    data = response.json()
    results = []

    for res in data.get("result", []):
        for series in res.get("data", []):
            service_id = series["dimensions"][0]
            values = [v for v in series.get("values", []) if v is not None]

            if not values:
                continue

            avg_value = sum(values) / len(values)

            results.append({
                "service_id": service_id,
                "service_name": get_service_name(tenant_url, api_token, service_id),
                "value": avg_value
            })

    return results


def get_service_name(tenant_url, api_token, service_id):
    url = f"{tenant_url.rstrip('/')}/api/v2/entities/{service_id}"

    headers = {
        "Authorization": f"Api-Token {api_token}",
        "Accept": "application/json"
    }

    try:
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            return (
                data.get("displayName")
                or data.get("properties", {}).get("serviceName")
                or service_id
            )
    except Exception:
        pass

    return service_id


######################Process Table ############################
from datetime import date
from django.http import JsonResponse
from .models import Process

TENANT_URL3 = "https://kdqxxx.live.dynatrace.com"
API_TOKEN3 = ""

def ProcessLevelMetrics(request):
    total_saved = 0

    for i in range(0, 5):  # last 7 days
        day = date.today() - timedelta(days=i)

        start_time = f"now-{i}d/d"
        end_time = f"now-{i-1}d/d" if i > 1 else "now/d"

        availability_data = query_process_metric(
            TENANT_URL3, API_TOKEN3,
            "builtin:pgi.availability",
            start_time, end_time
        )

        cpu_data = query_process_metric(
            TENANT_URL3, API_TOKEN3,
            "builtin:tech.generic.cpu.usage",
            start_time, end_time
        )

        memory_data = query_process_metric(
            TENANT_URL3, API_TOKEN3,
            "builtin:tech.generic.mem.usage",
            start_time, end_time
        )

        all_processes = set(availability_data) | set(cpu_data) | set(memory_data)

        for process_id in all_processes:
            process_name = (
                availability_data.get(process_id, {}).get("process_name")
                or cpu_data.get(process_id, {}).get("process_name")
                or memory_data.get(process_id, {}).get("process_name")
                or process_id
            )

            Process.objects.update_or_create(
                process_name=process_name,
                record_date=day,
                defaults={
                    "availability": round(availability_data.get(process_id, {}).get("value", 0), 2),
                    "cpu_usage": round(cpu_data.get(process_id, {}).get("value", 0), 2),
                    "memory_usage": round(memory_data.get(process_id, {}).get("value", 0), 2),
                }
            )

            total_saved += 1

    return JsonResponse({"status": "success", "processes_saved": total_saved})
import requests

def query_process_metric(tenant_url, api_token, metric_key, start_time, end_time):
    url = f"{tenant_url.rstrip('/')}/api/v2/metrics/query"

    params = {
        "metricSelector": f'{metric_key}:splitBy("dt.entity.process_group_instance")',
        "from": start_time,
        "to": end_time,
        "resolution": "1h",
        "pageSize": 4000
    }

    headers = {
        "Authorization": f"Api-Token {api_token}",
        "Accept": "application/json"
    }

    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()

    data = response.json()
    results = {}

    for res in data.get("result", []):
        for series in res.get("data", []):
            process_id = series["dimensions"][0]
            values = [v for v in series.get("values", []) if v is not None]

            if not values:
                continue

            avg_value = sum(values) / len(values)

            results[process_id] = {
                "process_name": get_process_name(tenant_url, api_token, process_id),
                "value": avg_value
            }

    return results

def get_process_name(tenant_url, api_token, process_id):
    url = f"{tenant_url.rstrip('/')}/api/v2/entities/{process_id}"

    headers = {
        "Authorization": f"Api-Token {api_token}",
        "Accept": "application/json"
    }

    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()

            return (
                data.get("displayName")
                or data.get("properties", {}).get("processGroupInstanceName")
                or data.get("properties", {}).get("processGroupName")
                or process_id
            )
    except Exception:
        pass

    return process_id

################Modal GRAPH FOR NEXT 7 DAYS PREDICTION##############
from django.http import JsonResponse
from datetime import date, timedelta
from django.db.models import Avg
from .models import Host, Process, Service

QUERY_MAP = {
    # Host
    "host cpu": {"model": Host, "field": "cpu_usage", "label": "CPU Usage"},
    "host memory": {"model": Host, "field": "memory_usage", "label": "Memory Usage"},
    "host availability": {"model": Host, "field": "host_availability", "label": "Availability"},
    # Process
    "process cpu": {"model": Process, "field": "cpu_usage", "label": "CPU Usage"},
    "process memory": {"model": Process, "field": "memory_usage", "label": "Memory Usage"},
    "process availability": {"model": Process, "field": "availability", "label": "Availability"},
    # Services
    "service request": {"model": Service, "field": "request_count", "label": "Request Count"},
    "service response": {"model": Service, "field": "response_time", "label": "Response Time"},
    "service failure": {"model": Service, "field": "failure_rate", "label": "Failure Rate"},
}

def AIQueryRouter(request):
    query = request.GET.get("q", "").lower()
    for key, config in QUERY_MAP.items():
        if all(word in query for word in key.split()):
            return forecast_json(config["model"], config["field"], config["label"])
    return JsonResponse({"error": "Query not understood"}, status=400)

from datetime import date, timedelta
from django.http import JsonResponse
from django.db.models import Avg


def forecast_json(model, field, label):
    today = date.today()

    last_record = model.objects.order_by("-record_date").first()
    if not last_record:
        return JsonResponse({"error": "No historical data found"}, status=400)

    last_data_date = last_record.record_date

    # =========================
    # 📉 GET LAST 60 DATA POINTS
    # =========================
    history_qs = (
        model.objects.filter(record_date__lte=last_data_date)
        .values("record_date")
        .annotate(avg_value=Avg(field))
        .order_by("-record_date")[:60]
    )

    history = list(history_qs)[::-1]

    if len(history) < 5:
        return JsonResponse({"error": "Not enough historical data"}, status=400)

    dates = [row["record_date"] for row in history]
    values = [float(row["avg_value"]) for row in history]

    # =========================
    # 1️⃣ DAY-TO-DAY TREND
    # =========================
    daily_deltas = [values[i] - values[i - 1] for i in range(1, len(values))]
    avg_daily_delta = sum(daily_deltas) / len(daily_deltas)

    # =========================
    # 2️⃣ SAME WEEKDAY TREND
    # =========================
    weekday_map = {}
    for d, v in zip(dates, values):
        weekday_map.setdefault(d.weekday(), []).append(v)

    weekday_deltas = []
    for vals in weekday_map.values():
        if len(vals) > 1:
            weekday_deltas.extend([vals[i] - vals[i - 1] for i in range(1, len(vals))])

    avg_weekday_delta = sum(weekday_deltas) / len(weekday_deltas) if weekday_deltas else 0

    # =========================
    # 3️⃣ MONTH START / END TREND
    # =========================
    month_edge_values = [v for d, v in zip(dates, values) if d.day <= 5 or d.day >= 25]

    month_edge_deltas = []
    if len(month_edge_values) > 1:
        month_edge_deltas = [
            month_edge_values[i] - month_edge_values[i - 1]
            for i in range(1, len(month_edge_values))
        ]

    avg_month_edge_delta = sum(month_edge_deltas) / len(month_edge_deltas) if month_edge_deltas else 0

    # =========================
    # 4️⃣ WEEKEND IMPACT
    # =========================
    weekend_values = [v for d, v in zip(dates, values) if d.weekday() >= 5]
    weekday_values = [v for d, v in zip(dates, values) if d.weekday() < 5]

    if weekend_values and weekday_values:
        avg_weekend = sum(weekend_values) / len(weekend_values)
        avg_weekday = sum(weekday_values) / len(weekday_values)
        weekend_impact_delta = avg_weekend - avg_weekday  # usually negative
    else:
        weekend_impact_delta = 0

    # =========================
    # 🔥 BASE DELTA COMBINATION
    # =========================
    base_delta = (
        avg_daily_delta * 0.4 +
        avg_weekday_delta * 0.25 +
        avg_month_edge_delta * 0.15
    )

    last_value = values[-1]

    # =========================
    # 📅 LAST 7 ACTUAL DATA DAYS
    # =========================
    last7_qs = (
        model.objects.filter(record_date__lte=last_data_date)
        .values("record_date")
        .annotate(avg_value=Avg(field))
        .order_by("-record_date")[:7]
    )

    last7_list = list(last7_qs)[::-1]

    actual_dates = [row["record_date"].strftime("%d-%m") for row in last7_list]
    actual_values = [round(float(row["avg_value"]), 2) for row in last7_list]

    # =========================
    # 🔮 FUTURE 7 DAY PREDICTION
    # =========================
    future_dates = []
    predicted_values = []

    prediction_start = today + timedelta(days=1)

    for i in range(7):
        next_day = prediction_start + timedelta(days=i)
        smart_delta = base_delta

        # Weekday behaviour
        if next_day.weekday() in weekday_map:
            smart_delta += avg_weekday_delta * 0.1

        # Month edge behaviour
        if next_day.day <= 5 or next_day.day >= 25:
            smart_delta += avg_month_edge_delta * 0.1

        # Weekend drop
        if next_day.weekday() == 5:  # Saturday
          smart_delta -= abs(base_delta) * 0.6
        elif next_day.weekday() == 6:  # Sunday
          smart_delta -= abs(base_delta) * 0.9

        next_val = last_value + smart_delta * (i + 1)

        metric = label.lower()
        if "availability" in metric:
            next_val = min(100, max(95, next_val))
        elif "cpu" in metric or "memory" in metric:
            next_val = min(100, max(0, next_val))
        else:
            next_val = max(0, next_val)

        future_dates.append(next_day.strftime("%d-%m"))
        predicted_values.append(round(next_val, 2))

    # =========================
    # 📊 FINAL GRAPH RESPONSE
    # =========================
    return JsonResponse({
        "metric": label,
        "labels": actual_dates + future_dates,
        "actual": actual_values + [None] * 7,
        "predicted": [None] * len(actual_values) + predicted_values
    })


# //////////////////////////////Generative AI /////////////////////////////////
import re
from datetime import timedelta
from django.http import JsonResponse
from .models import Host, Process, Service


def ai_search(request):
    raw_query = request.GET.get("q", "")
    query = raw_query.lower()

    # ---------------- EXTRACT DAYS FIRST ----------------
    days_match = re.search(r"(\d+)\s*day", query)
    days = int(days_match.group(1)) if days_match else None

    # ---------------- EXTRACT METRIC VALUE ----------------
    value = None

    # Extract number anywhere in query
    number_match = re.search(r"(\d+(\.\d+)?)", query)
    if number_match:
        value = float(number_match.group(1))


    def is_greater(q):
        return any(word in q for word in [
            ">", "greater", "greater than", "above", "more than", "over"
        ])

    def is_less(q):
        return any(word in q for word in [
            "<", "less", "less than", "below", "under"
        ])

    labels = []
    values = []
    metric = None
    source = None

    # ================== SOURCE SELECTION ==================
    if "host" in query:
        source = "host"
        qs = Host.objects.all()

    elif "process" in query:
        source = "process"
        qs = Process.objects.all()

    elif "service" in query:
        source = "service"
        qs = Service.objects.all()

    else:
        return JsonResponse({"error": "Please specify host, process or service"})

    # ================== DATE FILTER FIRST ==================
    qs = qs.order_by("-record_date")
    latest = qs.first()

    if latest:
        if days:
            end_date = latest.record_date
            start_date = end_date - timedelta(days=days)
            qs = qs.filter(record_date__range=[start_date, end_date])
        else:
            qs = qs.filter(record_date=latest.record_date)

    qs = qs.order_by("record_date")

    # ================== METRIC FILTER AFTER DATE ==================
    if source in ["host", "process"]:

        if "cpu" in query:
            metric = "cpu_usage"
        elif "memory" in query:
            metric = "memory_usage"
        elif "availability" in query:
            metric = "host_availability" if source == "host" else "availability"

    elif source == "service":

        if "response" in query or "time" in query:
            metric = "response_time"
        elif "failure" in query:
            metric = "failure_rate"
        elif "request" in query:
            metric = "request_count"

    # -------- APPLY FILTER --------
    if metric and value is not None:
        if is_greater(query):
            qs = qs.filter(**{f"{metric}__gt": value})
        elif is_less(query):
            qs = qs.filter(**{f"{metric}__lt": value})


    # ================== BUILD RESPONSE ==================
    for obj in qs:
        labels.append(obj.record_date.strftime("%d %b"))
        values.append(getattr(obj, metric) if metric else 0)

    results = list(qs.values())

    return JsonResponse({
        "query": raw_query,
        "source": source,
        "metric": metric,
        "count": len(results),
        "results": results,
        "chart": {
            "labels": labels,
            "values": values
        }
    })

# ====================================================================================
# ✅ END OF FILE
# ====================================================================================

